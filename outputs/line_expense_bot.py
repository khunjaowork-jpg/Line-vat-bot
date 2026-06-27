from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import hmac
import io
import json
import mimetypes
import os
import re
import sys
import uuid
import urllib.parse
import urllib.request
from urllib.parse import unquote
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

CONFIG: dict[str, Any] = {}
CONFIG_PATH: Path | None = None
STATE_CACHE: dict[str, Any] = {}


def runtime_log(message: str) -> None:
    timestamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line, file=sys.stderr)
    try:
        Path("outputs/line_bot_runtime.log").open("a", encoding="utf-8").write(line + "\n")
    except Exception:
        pass


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    if os.getenv("GOOGLE_VISION_API_KEY"):
        config["google_vision_api_key"] = os.getenv("GOOGLE_VISION_API_KEY")
    if os.getenv("GOOGLE_APPS_SCRIPT_URL"):
        config["google_apps_script_url"] = os.getenv("GOOGLE_APPS_SCRIPT_URL")
    if os.getenv("GOOGLE_APPS_SCRIPT_SECRET"):
        config["google_apps_script_secret"] = os.getenv("GOOGLE_APPS_SCRIPT_SECRET")
    if os.getenv("VAT_RATE"):
        config["vat_rate"] = float(os.getenv("VAT_RATE", "0.07"))
    if os.getenv("PORT"):
        config["port"] = int(os.getenv("PORT", "8080"))
    if os.getenv("TESSERACT_CMD"):
        config["tesseract_cmd"] = os.getenv("TESSERACT_CMD")
    if os.getenv("TESSDATA_DIR"):
        config["tessdata_dir"] = os.getenv("TESSDATA_DIR")
    return config


def resolve_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if path.is_absolute():
        return path
    if CONFIG_PATH is None:
        return path.resolve()
    return (CONFIG_PATH.parent.parent / path).resolve()


def state_path() -> Path:
    return resolve_path(CONFIG.get("state_file", "outputs/line_bot_state.json"))


def reply_image_dir() -> Path:
    return resolve_path(CONFIG.get("reply_image_dir", "outputs/line_reply_images"))


def load_state_cache() -> dict[str, Any]:
    path = state_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception as exc:
        runtime_log(f"State load failed: {exc}")
        return {}


def save_state_cache() -> None:
    path = state_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(STATE_CACHE, handle, ensure_ascii=False, indent=2, default=str)


def get_user_state(line_user_id: str) -> dict[str, Any]:
    if not line_user_id:
        line_user_id = "unknown"
    return dict(STATE_CACHE.get(line_user_id, {}))


def set_user_state(line_user_id: str, state: dict[str, Any]) -> None:
    if not line_user_id:
        line_user_id = "unknown"
    state["updated_at"] = dt.datetime.now().isoformat(timespec="seconds")
    STATE_CACHE[line_user_id] = state
    save_state_cache()


def clear_user_state(line_user_id: str) -> None:
    if not line_user_id:
        line_user_id = "unknown"
    if line_user_id in STATE_CACHE:
        del STATE_CACHE[line_user_id]
        save_state_cache()


def serialize_data(data: dict[str, Any]) -> dict[str, Any]:
    serialized = dict(data)
    if isinstance(serialized.get("date"), (dt.date, dt.datetime)):
        serialized["date"] = serialized["date"].isoformat()
    return serialized


def deserialize_data(data: dict[str, Any]) -> dict[str, Any]:
    restored = dict(data)
    if isinstance(restored.get("date"), str):
        try:
            restored["date"] = dt.date.fromisoformat(restored["date"][:10])
        except ValueError:
            restored["date"] = dt.date.today()
    return restored


def verify_line_signature(body: bytes, signature: str, channel_secret: str) -> bool:
    digest = hmac.new(channel_secret.encode("utf-8"), body, hashlib.sha256).digest()
    expected = base64.b64encode(digest).decode("utf-8")
    return hmac.compare_digest(expected, signature or "")


def line_request(method: str, url: str, token: str, data: bytes | None = None, content_type: str | None = None) -> bytes:
    headers = {"Authorization": f"Bearer {token}"}
    if content_type:
        headers["Content-Type"] = content_type
    request = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def post_json(url: str, payload: dict[str, Any]) -> dict[str, Any]:
    data = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        body = response.read().decode("utf-8")
    try:
        return json.loads(body)
    except json.JSONDecodeError:
        return {"status": "raw", "body": body}


def download_line_content(message_id: str, token: str, archive_dir: Path) -> Path:
    archive_dir.mkdir(parents=True, exist_ok=True)
    url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
    data = line_request("GET", url, token)
    path = archive_dir / f"{dt.datetime.now():%Y%m%d_%H%M%S}_{message_id}_{uuid.uuid4().hex[:8]}.jpg"
    path.write_bytes(data)
    return path


def get_line_display_name(line_user_id: str) -> str:
    if not line_user_id:
        return ""
    line_config = CONFIG.get("line", {})
    token_name = line_config.get("channel_access_token_env", "LINE_CHANNEL_ACCESS_TOKEN")
    token = os.getenv(token_name)
    if not token:
        return ""
    try:
        data = line_request("GET", f"https://api.line.me/v2/bot/profile/{line_user_id}", token)
        profile = json.loads(data.decode("utf-8"))
        return str(profile.get("displayName") or "").strip()
    except Exception as exc:
        runtime_log(f"LINE profile lookup failed: {exc}")
        return ""


def reply_messages(reply_token: str, messages: list[dict[str, Any]]) -> None:
    line_config = CONFIG["line"]
    if not line_config.get("reply_enabled", True):
        return
    token = os.getenv(line_config["channel_access_token_env"])
    if not token:
        runtime_log("LINE reply skipped: missing channel access token")
        return
    payload = {"replyToken": reply_token, "messages": messages[:5]}
    try:
        line_request(
            "POST",
            "https://api.line.me/v2/bot/message/reply",
            token,
            data=json.dumps(payload).encode("utf-8"),
            content_type="application/json",
        )
        runtime_log("LINE reply sent")
    except Exception as exc:
        runtime_log(f"LINE reply failed: {exc}")


def push_line_messages(to_id: str, messages: list[dict[str, Any]]) -> None:
    line_config = CONFIG.get("line", {})
    token_name = line_config.get("channel_access_token_env", "LINE_CHANNEL_ACCESS_TOKEN")
    token = os.getenv(token_name)
    if not token or not to_id:
        runtime_log("LINE push skipped: missing token or target id")
        return
    payload = {"to": to_id, "messages": messages[:5]}
    try:
        line_request(
            "POST",
            "https://api.line.me/v2/bot/message/push",
            token,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            content_type="application/json",
        )
        runtime_log(f"LINE push sent to {to_id}")
    except Exception as exc:
        runtime_log(f"LINE push failed: {exc}")


def text_message(text: str) -> dict[str, Any]:
    return {"type": "text", "text": text[:4900]}


def quick_reply_text_message(text: str, buttons: list[tuple[str, str]]) -> dict[str, Any]:
    message = text_message(text)
    message["quickReply"] = {
        "items": [
            {
                "type": "action",
                "action": {
                    "type": "message",
                    "label": label[:20],
                    "text": value,
                },
            }
            for label, value in buttons
        ][:13]
    }
    return message


def buttons_template_message(text: str, buttons: list[tuple[str, str]], title: str | None = None) -> dict[str, Any]:
    template: dict[str, Any] = {
        "type": "buttons",
        "text": text[:160],
        "actions": [
            {
                "type": "message",
                "label": label[:20],
                "text": value,
            }
            for label, value in buttons
        ][:4],
    }
    if title:
        template["title"] = title[:40]
    return {
        "type": "template",
        "altText": text[:400],
        "template": template,
    }


def image_message(public_url: str) -> dict[str, Any]:
    return {
        "type": "image",
        "originalContentUrl": public_url,
        "previewImageUrl": public_url,
    }


def reply_text(reply_token: str, text: str) -> None:
    reply_messages(reply_token, [text_message(text)])


def normalize_amount(value: str) -> float | None:
    cleaned = re.sub(r"[^\d.,-]", "", value).replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_date(text: str) -> dt.date | None:
    patterns = [
        r"(?<!\d)(\d{4})[/-](\d{1,2})[/-](\d{1,2})(?!\d)",
        r"(?<!\d)(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})(?!\d)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        parts = [int(part) for part in match.groups()]
        try:
            if len(str(parts[0])) == 4:
                year, month, day = parts
                if year > 2400:
                    year -= 543
            else:
                day, month, year = parts
                if year < 100:
                    year += 2000
                if year > 2400:
                    year -= 543
            return dt.date(year, month, day)
        except ValueError:
            continue
    return None


def find_amount_after_keywords(text: str, keywords: list[str]) -> float | None:
    for keyword in keywords:
        pattern = rf"{keyword}[\s:]*([0-9,]+(?:\.\d{{1,2}})?)"
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            amount = normalize_amount(match.group(1))
            if amount is not None:
                return amount
    return None


def find_likely_total(text: str) -> float | None:
    amounts: list[float] = []
    for match in re.finditer(r"(?<![\d-])(\d{1,3}(?:,\d{3})*|\d+)\.\d{2}(?!\d)", text):
        amount = normalize_amount(match.group(0))
        if amount is not None and 0 < amount < 10_000_000:
            amounts.append(amount)
    if not amounts:
        return None
    return max(amounts)


def extract_document_no(text: str) -> str:
    keywords = [
        "invoice no",
        "invoice no.",
        "invoice number",
        "tax invoice no",
        "tax invoice no.",
        "receipt no",
        "receipt no.",
        "receipt number",
        "document no",
        "document number",
        "เน€เธฅเธเธ—เธตเนเน€เธญเธเธชเธฒเธฃ",
        "เน€เธฅเธเธ—เธตเนเธเธดเธฅ",
        "เน€เธฅเธเธ—เธตเนเนเธเน€เธชเธฃเนเธ",
        "เน€เธฅเธเธ—เธตเนเนเธเธเธณเธเธฑเธ",
        "เน€เธฅเธเธ—เธตเน",
    ]
    # Match line-by-line first so the captured value does not drift into later text.
    candidates = [line.strip() for line in text.splitlines() if line.strip()]
    candidates.append(re.sub(r"\s+", " ", text))

    value_pattern = r"([A-Z0-9เธ-เน][A-Z0-9เธ-เน./_-]{1,40})"
    for candidate in candidates:
        for keyword in keywords:
            keyword_pattern = re.escape(keyword).replace(r"\ ", r"\s*")
            pattern = rf"{keyword_pattern}\s*(?:[:#\-]|no\.?|number|เน€เธฅเธเธ—เธตเน)?\s*{value_pattern}"
            match = re.search(pattern, candidate, flags=re.IGNORECASE)
            if not match:
                continue
            value = match.group(1).strip(" .:-#")
            if value and value.lower() not in {"date", "no", "number"}:
                return value
    return ""


def normalize_invoice_no(value: Any) -> str:
    text = str(value or "").strip()
    return text if text else "-"


def extract_document_type(text: str) -> str:
    compact = re.sub(r"\s+", " ", text).lower()
    thai_compact = re.sub(r"\s+", "", text)
    if "tax invoice" in compact or "เนเธเธเธณเธเธฑเธเธ เธฒเธฉเธต" in thai_compact or "เนเธเธเนเธฒเธเธฑเธเธ เธฒเธฉเธต" in thai_compact:
        return "เนเธเธเธณเธเธฑเธเธ เธฒเธฉเธต"
    if "receipt" in compact or "เนเธเน€เธชเธฃเนเธ" in thai_compact:
        return "เนเธเน€เธชเธฃเนเธ"
    if "bill" in compact or "เธเธดเธฅ" in thai_compact:
        return "เธเธดเธฅ"
    if "invoice" in compact:
        return "เนเธเธเธณเธเธฑเธเธ เธฒเธฉเธต"
    return "เธเธดเธฅ/เนเธเน€เธชเธฃเนเธ"


def parse_receipt_text(text: str, vat_rate: float) -> dict[str, Any]:
    compact = re.sub(r"\s+", " ", text)
    total = find_amount_after_keywords(
        compact,
        ["เธขเธญเธ”เธฃเธงเธก", "เธฃเธงเธกเธ—เธฑเนเธเธชเธดเนเธ", "เธเธณเธเธงเธเน€เธเธดเธเธฃเธงเธก", "total", "grand total", "amount due"],
    )
    vat = find_amount_after_keywords(
        compact,
        ["เธ เธฒเธฉเธตเธกเธนเธฅเธเนเธฒเน€เธเธดเนเธก", "เธ เธฒเธฉเธต", "vat", "value added tax"],
    )
    withholding_tax = find_amount_after_keywords(
        compact,
        ["withholding tax", "wht", "tax withheld", "\u0e20\u0e32\u0e29\u0e35\u0e2b\u0e31\u0e01 \u0e13 \u0e17\u0e35\u0e48\u0e08\u0e48\u0e32\u0e22", "\u0e20\u0e32\u0e29\u0e35\u0e2b\u0e31\u0e01\u0e13\u0e17\u0e35\u0e48\u0e08\u0e48\u0e32\u0e22"],
    )
    before_vat = find_amount_after_keywords(
        compact,
        ["เธขเธญเธ”เธเนเธญเธเธ เธฒเธฉเธต", "เธกเธนเธฅเธเนเธฒเธชเธดเธเธเนเธฒ", "subtotal", "before vat"],
    )

    if total is None:
        total = find_likely_total(compact)

    if before_vat is None and total is not None:
        if vat is not None:
            before_vat = total - vat
        else:
            before_vat = round(total / (1 + vat_rate), 2)
            vat = round(total - before_vat, 2)
    if vat is None and before_vat is not None:
        vat = round(before_vat * vat_rate, 2)
    if total is None and before_vat is not None:
        total = round(before_vat + (vat or 0), 2)

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    vendor = lines[0][:80] if lines else ""
    invoice_no = normalize_invoice_no(extract_document_no(text))

    confidence = 0.25
    if parse_date(compact):
        confidence += 0.2
    if vendor:
        confidence += 0.15
    if before_vat is not None:
        confidence += 0.2
    if vat is not None:
        confidence += 0.1
    if total is not None:
        confidence += 0.1

    return {
        "date": parse_date(compact) or dt.date.today(),
        "document_type": extract_document_type(text),
        "invoice_no": invoice_no,
        "vendor": vendor,
        "description": "LINE receipt OCR",
        "category": CONFIG.get("default_category", "เธญเธทเนเธ เน"),
        "before_vat": round(before_vat or 0, 2),
        "vat": round(vat or 0, 2),
        "withholding_tax": round(withholding_tax or 0, 2),
        "total": round(total or 0, 2),
        "claimable": CONFIG.get("default_claimable", "Yes") if (vat or 0) > 0 else "No",
        "confidence": min(confidence, 1.0),
        "raw_text": compact[:500],
    }


def apply_transaction_type_defaults(data: dict[str, Any], transaction_type: str) -> dict[str, Any]:
    updated = dict(data)
    normalized_type = "Revenue" if str(transaction_type).lower() == "revenue" else "Expense"
    updated["transaction_type"] = normalized_type
    updated["invoice_no"] = normalize_invoice_no(updated.get("invoice_no"))
    if normalized_type == "Revenue":
        updated["category"] = updated.get("category") or CONFIG.get("default_revenue_category", "Revenue")
        updated["claimable"] = "Yes"
        if updated.get("description") == "LINE receipt OCR":
            updated["description"] = "LINE revenue OCR"
    else:
        updated["category"] = updated.get("category") or CONFIG.get("default_category", "Other")
        updated["claimable"] = updated.get("claimable") or CONFIG.get("default_claimable", "Yes")
    return updated


def ocr_image(image_path: Path) -> str:
    vision_api_key = CONFIG.get("google_vision_api_key") or os.getenv(CONFIG.get("google_vision_api_key_env", "GOOGLE_VISION_API_KEY"))
    if vision_api_key:
        try:
            runtime_log("Google Vision OCR started")
            text = google_vision_ocr_image(image_path, vision_api_key)
            runtime_log(f"Google Vision OCR completed characters={len(text)}")
            if text.strip():
                return text
            raise RuntimeError("Google Vision OCR returned empty text")
        except Exception as exc:
            raise RuntimeError(f"Google Vision OCR failed within time limit: {exc}") from exc

    try:
        from PIL import Image, ImageOps
        import pytesseract
    except ImportError as exc:
        raise RuntimeError("Install OCR packages first: pip install pillow pytesseract") from exc

    lang = CONFIG.get("tesseract_lang", "tha+eng")
    if CONFIG.get("tesseract_cmd"):
        tesseract_cmd = str(CONFIG["tesseract_cmd"])
        if Path(tesseract_cmd).is_absolute() or "/" in tesseract_cmd or "\\" in tesseract_cmd:
            tesseract_cmd = str(resolve_path(tesseract_cmd))
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    if CONFIG.get("tessdata_dir"):
        tessdata_dir = resolve_path(CONFIG["tessdata_dir"])
        os.environ["TESSDATA_PREFIX"] = str(tessdata_dir)
    with Image.open(image_path) as image:
        image = ImageOps.exif_transpose(image)
        image.thumbnail((int(CONFIG.get("ocr_max_dimension", 1600)), int(CONFIG.get("ocr_max_dimension", 1600))))
        image = ImageOps.autocontrast(image.convert("L"))
        tesseract_config = CONFIG.get("tesseract_config", "--oem 1 --psm 11")
        timeout = int(CONFIG.get("ocr_timeout_seconds", 15))
        try:
            return pytesseract.image_to_string(image, lang=lang, config=tesseract_config, timeout=timeout)
        except RuntimeError as exc:
            raise RuntimeError(f"Tesseract OCR failed within time limit: {exc}") from exc


def google_vision_ocr_image(image_path: Path, api_key: str) -> str:
    image_content = base64.b64encode(image_path.read_bytes()).decode("ascii")
    payload = {
        "requests": [
            {
                "image": {"content": image_content},
                "features": [{"type": "DOCUMENT_TEXT_DETECTION"}],
                "imageContext": {"languageHints": ["th", "en"]},
            }
        ]
    }
    request = urllib.request.Request(
        f"https://vision.googleapis.com/v1/images:annotate?key={api_key}",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    timeout = int(CONFIG.get("google_vision_timeout_seconds", 15))
    with urllib.request.urlopen(request, timeout=timeout) as response:
        result = json.loads(response.read().decode("utf-8"))

    if result.get("error"):
        raise RuntimeError(result["error"].get("message", result["error"]))

    responses = result.get("responses", [])
    if not responses:
        return ""
    first = responses[0]
    if first.get("error"):
        raise RuntimeError(first["error"].get("message", first["error"]))
    if first.get("fullTextAnnotation", {}).get("text"):
        return first["fullTextAnnotation"]["text"]
    annotations = first.get("textAnnotations") or []
    if annotations:
        return annotations[0].get("description", "")
    return ""


def next_empty_row(sheet, start_row: int = 4) -> int:
    row = start_row
    while sheet.cell(row=row, column=1).value:
        row += 1
    return row


def copy_formula_row(sheet, row: int, cols: tuple[int, ...]) -> None:
    if row <= 4:
        return
    for col in cols:
        source = sheet.cell(row=row - 1, column=col)
        target = sheet.cell(row=row, column=col)
        if isinstance(source.value, str) and source.value.startswith("="):
            target.value = source.value.replace(str(row - 1), str(row))


def find_transaction_row(sheet, date_value: dt.date) -> int:
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    target_month = month_names[date_value.month - 1]
    block_start = None

    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        if isinstance(value, str) and target_month in value:
            block_start = row + 2
            break
    if block_start is None:
        raise RuntimeError("Cannot find matching month block in Transactions_12M.")

    block_end = min(block_start + 29, sheet.max_row)
    for row in range(block_start, block_end + 1):
        if not sheet.cell(row=row, column=1).value:
            return row
    raise RuntimeError(f"No empty rows left in Transactions_12M for {date_value:%Y-%m}.")


def append_to_transactions(sheet, row: int, image_path: Path, data: dict[str, Any]) -> None:
    sheet.cell(row=row, column=1, value=data["date"])
    sheet.cell(row=row, column=2, value=data.get("transaction_type", "Expense"))
    sheet.cell(row=row, column=3, value=normalize_invoice_no(data.get("invoice_no")))
    sheet.cell(row=row, column=4, value=data["vendor"])
    sheet.cell(row=row, column=5, value=data["description"])
    sheet.cell(row=row, column=6, value=data["category"])
    sheet.cell(row=row, column=7, value=data["before_vat"])
    sheet.cell(row=row, column=8, value=CONFIG.get("vat_rate", 0.07))
    sheet.cell(row=row, column=9, value=f'=IF(G{row}="","",IF(OR(B{row}="Revenue",K{row}="Yes"),ROUND(G{row}*H{row},2),0))')
    sheet.cell(row=row, column=10, value=f'=IF(G{row}="","",G{row}+ROUND(G{row}*H{row},2))')
    sheet.cell(row=row, column=11, value=data["claimable"])
    sheet.cell(row=row, column=12, value=f'=IF(A{row}="","",DATE(YEAR(A{row}),MONTH(A{row}),1))')
    sheet.cell(row=row, column=13, value=str(image_path))
    sheet.cell(row=row, column=14, value=data["confidence"])
    sheet.cell(row=row, column=15, value=f'=IF(B{row}="Revenue",G{row},0)')
    sheet.cell(row=row, column=16, value=f'=IF(B{row}="Expense",G{row},0)')
    sheet.cell(row=row, column=17, value=data["raw_text"])
    sheet.cell(row=row, column=18, value=data.get("document_type", "เธเธดเธฅ/เนเธเน€เธชเธฃเนเธ"))


def append_to_legacy_expenses(sheet, row: int, image_path: Path, data: dict[str, Any]) -> None:
    sheet.cell(row=row, column=1, value=data["date"])
    sheet.cell(row=row, column=2, value=normalize_invoice_no(data.get("invoice_no")))
    sheet.cell(row=row, column=3, value=data["vendor"])
    sheet.cell(row=row, column=4, value=data["description"])
    sheet.cell(row=row, column=5, value=data["category"])
    sheet.cell(row=row, column=6, value=data["before_vat"])
    sheet.cell(row=row, column=7, value=CONFIG.get("vat_rate", 0.07))
    sheet.cell(row=row, column=10, value=data["claimable"])
    sheet.cell(row=row, column=12, value=str(image_path))
    sheet.cell(row=row, column=13, value=data["confidence"])
    sheet.cell(row=row, column=14, value=data["raw_text"])
    sheet.cell(row=row, column=15, value=data.get("document_type", "เธเธดเธฅ/เนเธเน€เธชเธฃเนเธ"))
    copy_formula_row(sheet, row, (8, 9, 11))


def append_to_legacy_revenue(sheet, row: int, data: dict[str, Any]) -> None:
    sheet.cell(row=row, column=1, value=data["date"])
    sheet.cell(row=row, column=2, value=normalize_invoice_no(data.get("invoice_no")))
    sheet.cell(row=row, column=3, value=data["vendor"])
    sheet.cell(row=row, column=4, value=data["description"])
    sheet.cell(row=row, column=5, value=data["before_vat"])
    sheet.cell(row=row, column=6, value=CONFIG.get("vat_rate", 0.07))
    sheet.cell(row=row, column=7, value=f'=IF(E{row}="","",ROUND(E{row}*F{row},2))')
    sheet.cell(row=row, column=8, value=f'=IF(E{row}="","",E{row}+G{row})')
    sheet.cell(row=row, column=9, value="Yes")
    sheet.cell(row=row, column=10, value=f'=IF(A{row}="","",DATE(YEAR(A{row}),MONTH(A{row}),1))')
    sheet.cell(row=row, column=11, value=data.get("document_type", ""))


def append_import_log(log, image_path: Path, data: dict[str, Any], status: str, message: str, line_user_id: str) -> None:
    log_row = next_empty_row(log)
    log.cell(row=log_row, column=1, value=dt.datetime.now())
    log.cell(row=log_row, column=2, value=str(image_path))
    log.cell(row=log_row, column=3, value=status)
    log.cell(row=log_row, column=4, value=data["date"])
    log.cell(row=log_row, column=5, value=data["vendor"])
    log.cell(row=log_row, column=6, value=data["total"])
    log.cell(row=log_row, column=7, value=data["vat"])
    log.cell(row=log_row, column=8, value=f"{message} LINE user: {line_user_id}")
    log.cell(row=log_row, column=9, value=data.get("document_type", ""))


def format_parsed_details(data: dict[str, Any], heading: str = "เธเธดเธฅเธเธณเน€เธเนเธฒ") -> str:
    return (
        f"==== {heading} ====\n"
        f"เธเธฃเธฐเน€เธ เธ—: {data.get('transaction_type', '-')}\n"
        f"เธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ: {data.get('document_type') or '-'}\n"
        f"เธงเธฑเธเธ—เธตเน: {data.get('date')}\n"
        f"เน€เธฅเธเธ—เธตเนเธเธดเธฅ: {data.get('invoice_no') or '-'}\n"
        f"เธเธทเนเธญเธฃเนเธฒเธ/เธเธนเนเธเนเธฒ: {data.get('vendor') or '-'}\n"
        f"เธเธนเนเธเธณเธชเนเธเน€เธญเธเธชเธฒเธฃ: {data.get('submitter_name') or '-'}\n"
        f"เธซเธกเธงเธ”: {data.get('category') or '-'}\n"
        f"เธขเธญเธ”เธเนเธญเธ VAT: {float(data.get('before_vat') or 0):,.2f}\n"
        f"VAT: {float(data.get('vat') or 0):,.2f}\n"
        f"เธ เธฒเธฉเธตเธซเธฑเธ เธ“ เธ—เธตเนเธเนเธฒเธข: {float(data.get('withholding_tax') or 0):,.2f}\n"
        f"เธขเธญเธ”เธฃเธงเธก: {float(data.get('total') or 0):,.2f}\n"
        f"เธเธงเธฒเธกเธกเธฑเนเธเนเธ OCR: {float(data.get('confidence') or 0):.0%}"
    )


def confirmation_prompt(data: dict[str, Any]) -> str:
    return (
        format_parsed_details(data, "เธเธดเธฅเธเธณเน€เธเนเธฒ") + "\n\n"
        "เธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเธชเธญเธเธเนเธญเธกเธนเธฅเธเนเธญเธเธเธฑเธเธ—เธถเธเธฅเธ Excel\n"
        "เธ–เนเธฒเธ–เธนเธเธ•เนเธญเธ เธเธดเธกเธเน: เธ•เธฃเธงเธเธชเธญเธเนเธฅเธฐเธขเธทเธเธขเธฑเธ\n"
        "เธ–เนเธฒเธ•เนเธญเธเธเธฒเธฃเนเธเนเนเธ เธเธดเธกเธเน: เนเธเนเนเธ"
    )


def correction_form(data: dict[str, Any]) -> str:
    return (
        "เธเธฃเธธเธ“เธฒเนเธเนเนเธเธเนเธญเธกเธนเธฅเนเธเนเธเธเธเธญเธฃเนเธกเธเธตเน เนเธฅเนเธงเธชเนเธเธเธฅเธฑเธเธกเธฒเนเธ”เนเน€เธฅเธขเธเนเธฐ\n\n"
        f"เธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ: {data.get('document_type') or ''}\n"
        f"เธงเธฑเธเธ—เธตเน: {data.get('date') or ''}\n"
        f"เน€เธฅเธเธ—เธตเนเธเธดเธฅ: {normalize_invoice_no(data.get('invoice_no'))}\n"
        f"เธเธทเนเธญเธฃเนเธฒเธ/เธเธนเนเธเนเธฒ: {data.get('vendor') or ''}\n"
        f"เธเธนเนเธเธณเธชเนเธเน€เธญเธเธชเธฒเธฃ: {data.get('submitter_name') or ''}\n"
        f"เธซเธกเธงเธ”: {data.get('category') or ''}\n"
        f"เธขเธญเธ”เธเนเธญเธ VAT: {float(data.get('before_vat') or 0):.2f}\n"
        f"VAT: {float(data.get('vat') or 0):.2f}\n"
        f"เธ เธฒเธฉเธตเธซเธฑเธ เธ“ เธ—เธตเนเธเนเธฒเธข: {float(data.get('withholding_tax') or 0):.2f}\n"
        f"เธขเธญเธ”เธฃเธงเธก: {float(data.get('total') or 0):.2f}"
    )


def menu_text() -> str:
    return (
        "กรุณาเลือกเมนู\n"
        "1. บิลรายรับ\n"
        "2. บิลรายจ่าย\n"
        "3. เรียกดูรายละเอียดบัญชี\n"
        "4. ยกเลิกการทำรายการ\n\n"
        "พิมพ์เลขเมนูที่ต้องการได้เลยค่ะ"
    )


def account_menu_button(number: str, title: str, text: str, background: str, accent: str) -> dict[str, Any]:
    return {
        "type": "box",
        "layout": "horizontal",
        "cornerRadius": "18px",
        "backgroundColor": background,
        "paddingAll": "16px",
        "spacing": "14px",
        "action": {"type": "message", "label": title[:20], "text": text},
        "contents": [
            {
                "type": "box",
                "layout": "vertical",
                "width": "54px",
                "height": "54px",
                "cornerRadius": "27px",
                "backgroundColor": accent,
                "alignItems": "center",
                "justifyContent": "center",
                "contents": [
                    {
                        "type": "text",
                        "text": number,
                        "weight": "bold",
                        "size": "xl",
                        "color": "#FFFFFF",
                        "align": "center",
                    }
                ],
            },
            {
                "type": "text",
                "text": f"{number}. {title}",
                "weight": "bold",
                "size": "xl",
                "color": "#111C4E",
                "wrap": True,
                "gravity": "center",
                "flex": 1,
            },
            {
                "type": "text",
                "text": ">",
                "weight": "bold",
                "size": "xxl",
                "color": accent,
                "align": "end",
                "gravity": "center",
                "flex": 0,
            },
        ],
    }


EXPENSE_CATEGORIES = [
    ("ค่าน้ำ", "#DBEAFE", "#3B82F6", "น้ำ"),
    ("ค่าไฟ", "#FEF3C7", "#F59E0B", "ไฟ"),
    ("อินเตอร์เน็ต", "#EDE9FE", "#8B5CF6", "WiFi"),
    ("เบิกน้ำมันรถ", "#D1FAE5", "#10B981", "รถ"),
    ("สินค้าเติมสต็อค", "#E0F2FE", "#2563EB", "สต็อค"),
    ("ค่าธรรมเนียม", "#FCE7F3", "#EC4899", "%"),
    ("อื่นๆ", "#F1F5F9", "#64748B", "..."),
]

REVENUE_CATEGORIES = [
    ("ยอดขายหน้าสาขา", "#DCFCE7", "#22C55E", "สาขา"),
    ("ยอดขายส่ง", "#DBEAFE", "#3B82F6", "ส่ง"),
    ("อื่นๆ", "#F1F5F9", "#64748B", "..."),
]

REVENUE_BRANCHES = [
    ("สี่แยก", "#DBEAFE", "#3B82F6", "1"),
    ("พัสดุสี่แยก", "#FEF3C7", "#F59E0B", "2"),
    ("ทะเล", "#CCFBF1", "#14B8A6", "3"),
    ("เขาใหญ่", "#DCFCE7", "#22C55E", "4"),
    ("ภูเก็ต", "#EDE9FE", "#8B5CF6", "5"),
]

REVENUE_PAYMENT_CHANNELS = [
    ("ยอดขายเงินสด", "#DCFCE7", "#22C55E", "เงิน"),
    ("ยอดเงินโอน", "#DBEAFE", "#3B82F6", "โอน"),
    ("ยอดบัตรเครดิต", "#FEF3C7", "#F59E0B", "บัตร"),
    ("อื่นๆ", "#F1F5F9", "#64748B", "..."),
]


def category_menu_item(transaction_type: str, label: str, bg: str, accent: str, icon: str) -> dict[str, Any]:
    return {
        "type": "box",
        "layout": "horizontal",
        "cornerRadius": "16px",
        "backgroundColor": "#FFFFFF",
        "paddingAll": "14px",
        "spacing": "14px",
        "action": {"type": "message", "label": label[:20], "text": f"ACCT_CATEGORY:{transaction_type}:{label}"},
        "contents": [
            {
                "type": "box",
                "layout": "vertical",
                "width": "48px",
                "height": "48px",
                "cornerRadius": "14px",
                "backgroundColor": bg,
                "alignItems": "center",
                "justifyContent": "center",
                "contents": [
                    {
                        "type": "text",
                        "text": icon,
                        "weight": "bold",
                        "size": "xl",
                        "color": accent,
                        "align": "center",
                    }
                ],
            },
            {
                "type": "text",
                "text": label,
                "weight": "bold",
                "size": "xl",
                "color": "#1F2A44",
                "wrap": True,
                "gravity": "center",
                "flex": 1,
            },
            {
                "type": "text",
                "text": ">",
                "weight": "bold",
                "size": "xxl",
                "color": accent,
                "align": "end",
                "gravity": "center",
                "flex": 0,
            },
        ],
    }


def category_menu_message(transaction_type: str) -> dict[str, Any]:
    is_revenue = transaction_type == "Revenue"
    categories = REVENUE_CATEGORIES if is_revenue else EXPENSE_CATEGORIES
    title = "กรุณาเลือกประเภทรายได้" if is_revenue else "กรุณาเลือกประเภทค่าใช้จ่าย"
    accent = "#8B5CF6" if is_revenue else "#3B82F6"
    return {
        "type": "flex",
        "altText": title,
        "contents": {
            "type": "bubble",
            "size": "giga",
            "body": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#FFFFFF",
                "paddingAll": "18px",
                "spacing": "12px",
                "contents": [
                    {
                        "type": "box",
                        "layout": "horizontal",
                        "cornerRadius": "18px",
                        "backgroundColor": "#F8FAFC",
                        "borderColor": accent,
                        "borderWidth": "2px",
                        "paddingAll": "16px",
                        "contents": [
                            {
                                "type": "text",
                                "text": title,
                                "weight": "bold",
                                "size": "xl",
                                "color": "#111C4E",
                                "wrap": True,
                            }
                        ],
                    },
                    {
                        "type": "box",
                        "layout": "vertical",
                        "cornerRadius": "18px",
                        "backgroundColor": "#FFFFFF",
                        "paddingAll": "8px",
                        "spacing": "8px",
                        "contents": [
                            category_menu_item(transaction_type, label, bg, item_accent, icon)
                            for label, bg, item_accent, icon in categories
                        ],
                    },
                ],
            },
        },
    }


def simple_choice_menu(title: str, action_prefix: str, items: list[tuple[str, str, str, str]], accent: str = "#2563EB") -> dict[str, Any]:
    return {
        "type": "flex",
        "altText": title,
        "contents": {
            "type": "bubble",
            "size": "giga",
            "body": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#FFFFFF",
                "paddingAll": "18px",
                "spacing": "12px",
                "contents": [
                    {
                        "type": "box",
                        "layout": "horizontal",
                        "cornerRadius": "18px",
                        "backgroundColor": "#F8FAFC",
                        "borderColor": accent,
                        "borderWidth": "2px",
                        "paddingAll": "16px",
                        "contents": [
                            {
                                "type": "text",
                                "text": title,
                                "weight": "bold",
                                "size": "xl",
                                "color": "#111C4E",
                                "wrap": True,
                            }
                        ],
                    },
                    {
                        "type": "box",
                        "layout": "vertical",
                        "cornerRadius": "18px",
                        "backgroundColor": "#FFFFFF",
                        "paddingAll": "8px",
                        "spacing": "8px",
                        "contents": [
                            category_menu_item("Revenue", label, bg, item_accent, icon) | {
                                "action": {"type": "message", "label": label[:20], "text": f"{action_prefix}:{label}"}
                            }
                            for label, bg, item_accent, icon in items
                        ],
                    },
                ],
            },
        },
    }


def revenue_branch_menu_message() -> dict[str, Any]:
    return simple_choice_menu("กรุณาเลือกสาขา", "REV_BRANCH", REVENUE_BRANCHES)


def revenue_payment_menu_message() -> dict[str, Any]:
    return simple_choice_menu("กรุณาเลือกช่องทางรายรับ", "REV_PAYMENT", REVENUE_PAYMENT_CHANNELS, "#8B5CF6")


def continue_document_menu_message() -> dict[str, Any]:
    return simple_choice_menu(
        "ต้องการส่งเอกสารในรายการเดิมต่อหรือไม่",
        "DOC_CONTINUE",
        [
            ("ส่งเอกสารต่อ", "#DCFCE7", "#22C55E", "1"),
            ("ตรวจสอบและยืนยัน", "#EDE9FE", "#8B5CF6", "2"),
        ],
        "#10B981",
    )


def ask_for_receipt_after_category(category: str) -> str:
    return (
        f"เลือกหมวดหมู่: {category}\n\n"
        "ส่งเอกสารเพื่อบันทึกรายละเอียดในระบบได้เลยค่ะ"
    )


def format_revenue_payment_summary(payments: dict[str, Any]) -> str:
    lines = []
    for name, amount in payments.items():
        try:
            value = float(amount or 0)
        except (TypeError, ValueError):
            value = 0
        if value:
            lines.append(f"- {name}: {value:,.2f}")
    return "\n".join(lines) if lines else "-"


def apply_revenue_payments_to_data(data: dict[str, Any], state: dict[str, Any]) -> dict[str, Any]:
    payments = dict(state.get("revenue_payments") or {})
    if not payments:
        return data
    total = round(sum(float(value or 0) for value in payments.values()), 2)
    branch = str(state.get("revenue_branch") or "")
    category = str(state.get("category") or "ยอดขายหน้าสาขา")
    updated = dict(data)
    updated["transaction_type"] = "Revenue"
    updated["category"] = f"{category} - {branch}" if branch else category
    updated["description"] = " / ".join(f"{name} {float(value or 0):,.2f}" for name, value in payments.items() if float(value or 0))
    updated["before_vat"] = total
    updated["vat"] = 0
    updated["total"] = total
    updated["claimable"] = "Yes"
    updated["revenue_branch"] = branch
    updated["revenue_payments"] = payments
    return updated


def append_document_path_to_state(state: dict[str, Any], image_path: Path) -> list[str]:
    paths = list(state.get("document_paths") or [])
    text_path = str(image_path)
    if text_path not in paths:
        paths.append(text_path)
    state["document_paths"] = paths
    return paths


def notify_accounting_import_approver(data: dict[str, Any], line_user_id: str) -> None:
    approver_id = str(CONFIG.get("hr_approver_line_id") or os.getenv("HR_APPROVER_LINE_ID") or "Ud260925c43fb0823fea42224a2929393")
    if not approver_id:
        return
    message = (
        "มีรายการบัญชีใหม่บันทึกเข้าระบบแล้ว\n\n"
        f"ประเภท: {data.get('transaction_type') or '-'}\n"
        f"หมวด: {data.get('category') or '-'}\n"
        f"ร้าน/คู่ค้า: {data.get('vendor') or '-'}\n"
        f"ยอดรวม: {float(data.get('total') or 0):,.2f}\n"
        f"ผู้นำส่ง: {data.get('submitter_name') or line_user_id or '-'}"
    )
    try:
        push_line_messages(approver_id, [text_message(message)])
    except Exception as exc:
        runtime_log(f"Accounting approver notification failed: {exc}")


def menu_message() -> dict[str, Any]:
    buttons = [
        account_menu_button("1", "บิลรายรับ", "1", "#F1F7FF", "#3B82F6"),
        account_menu_button("2", "บิลรายจ่าย", "2", "#EFFCF8", "#10B981"),
        account_menu_button("3", "เรียกดูบัญชี", "3", "#FFF7E8", "#F59E0B"),
        account_menu_button("4", "ยกเลิกรายการ", "4", "#FFF1F6", "#EC4899"),
    ]
    return {
        "type": "flex",
        "altText": "กรุณาเลือกเมนูบัญชี",
        "contents": {
            "type": "bubble",
            "size": "giga",
            "body": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#FFFFFF",
                "paddingAll": "18px",
                "spacing": "14px",
                "contents": [
                    {
                        "type": "box",
                        "layout": "horizontal",
                        "cornerRadius": "18px",
                        "backgroundColor": "#F6F0FF",
                        "paddingAll": "16px",
                        "spacing": "12px",
                        "contents": [
                            {
                                "type": "box",
                                "layout": "vertical",
                                "width": "48px",
                                "height": "48px",
                                "cornerRadius": "24px",
                                "backgroundColor": "#7C3AED",
                                "alignItems": "center",
                                "justifyContent": "center",
                                "contents": [
                                    {
                                        "type": "text",
                                        "text": "เมนู",
                                        "weight": "bold",
                                        "size": "md",
                                        "color": "#FFFFFF",
                                        "align": "center",
                                    }
                                ],
                            },
                            {
                                "type": "text",
                                "text": "กรุณาเลือกเมนู",
                                "weight": "bold",
                                "size": "xxl",
                                "color": "#111C4E",
                                "gravity": "center",
                                "wrap": True,
                                "flex": 1,
                            },
                        ],
                    },
                    *buttons,
                ],
            },
        },
    }


def coming_soon_message(section: str) -> dict[str, Any]:
    return buttons_template_message(
        f"เธซเธกเธงเธ” {section}\n\n"
        "เธฃเธฐเธเธเธซเธกเธงเธ”เธเธตเนเธขเธฑเธเธญเธขเธนเนเธฃเธฐเธซเธงเนเธฒเธเน€เธ•เธฃเธตเธขเธกเนเธเนเธเธฒเธเธเนเธฐ\n"
        "เธ•เธญเธเธเธตเนเธชเธฒเธกเธฒเธฃเธ–เนเธเนเธเธฒเธเธซเธกเธงเธ”เธเธฑเธเธเธตเนเธ”เนเธเนเธญเธ",
        [
            ("เน€เธเธดเธ”เน€เธกเธเธนเธเธฑเธเธเธต", "เธเธฑเธเธเธต"),
        ],
    )


def stock_menu_message() -> dict[str, Any]:
    return stock_branch_menu_message()


def stock_branch_menu_message() -> dict[str, Any]:
    return buttons_template_message(
        "เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเธชเธฒเธเธฒเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธ•เธฃเธงเธเธชเธญเธเธชเธ•เนเธญเธ\nเธซเธฅเธฑเธเน€เธฅเธทเธญเธเธชเธฒเธเธฒ เธชเธฒเธกเธฒเธฃเธ–เธเธดเธกเธเนเธเธทเนเธญเธชเธดเธเธเนเธฒ/เธเธฒเธฃเนเนเธเนเธ” เธซเธฃเธทเธญเธชเนเธเธเธเธฒเธฃเนเนเธเนเธ”เนเธ”เนเธชเธนเธเธชเธธเธ” 10 เธฃเธฒเธขเธเธฒเธฃเธ•เนเธญเธเธฃเธฑเนเธ",
        [
            ("1. เธชเธตเนเนเธขเธ", "เธเนเธเธซเธฒเธชเธ•เนเธญเธ:เธชเธตเนเนเธขเธ"),
            ("2. เธเธฑเธชเธ”เธธเธชเธตเนเนเธขเธ", "เธเนเธเธซเธฒเธชเธ•เนเธญเธ:เธเธฑเธชเธ”เธธเธชเธตเนเนเธขเธ"),
            ("3. เธ—เธฐเน€เธฅ", "เธเนเธเธซเธฒเธชเธ•เนเธญเธ:เธ—เธฐเน€เธฅ"),
            ("4. เน€เธเธฒเนเธซเธเน", "เธเนเธเธซเธฒเธชเธ•เนเธญเธ:เน€เธเธฒเนเธซเธเน"),
        ],
    )

def schedule_month_menu_message() -> dict[str, Any]:
    return buttons_template_message(
        "เน€เธฅเธทเธญเธเน€เธ”เธทเธญเธเธ•เธฒเธฃเธฒเธเธเธฒเธเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธ”เธน",
        [
            ("เน€เธ”เธทเธญเธเธเนเธญเธเธซเธเนเธฒ", "เธ•เธฒเธฃเธฒเธเธเธฒเธ:-1"),
            ("เน€เธ”เธทเธญเธเธเธตเน", "เธ•เธฒเธฃเธฒเธเธเธฒเธ:0"),
            ("เน€เธ”เธทเธญเธเธ–เธฑเธ”เนเธ", "เธ•เธฒเธฃเธฒเธเธเธฒเธ:1"),
        ],
        title="เธ•เธฒเธฃเธฒเธเธเธฒเธ",
    )


def hr_menu_button(
    number: str,
    title: str,
    subtitle: str,
    text: str,
    background: str,
    accent: str,
) -> dict[str, Any]:
    return {
        "type": "box",
        "layout": "horizontal",
        "cornerRadius": "18px",
        "backgroundColor": background,
        "paddingAll": "14px",
        "spacing": "12px",
        "action": {"type": "message", "label": title[:20], "text": text},
        "contents": [
            {
                "type": "box",
                "layout": "vertical",
                "width": "50px",
                "height": "50px",
                "cornerRadius": "18px",
                "backgroundColor": accent,
                "alignItems": "center",
                "justifyContent": "center",
                "contents": [
                    {
                        "type": "text",
                        "text": number,
                        "weight": "bold",
                        "size": "xxl",
                        "color": "#FFFFFF",
                        "align": "center",
                    }
                ],
            },
            {
                "type": "box",
                "layout": "vertical",
                "flex": 1,
                "spacing": "3px",
                "justifyContent": "center",
                "contents": [
                    {
                        "type": "text",
                        "text": title,
                        "weight": "bold",
                        "size": "lg",
                        "color": "#1F2937",
                        "wrap": True,
                    },
                    {
                        "type": "text",
                        "text": subtitle,
                        "size": "sm",
                        "color": "#64748B",
                        "wrap": True,
                    },
                ],
            },
            {
                "type": "box",
                "layout": "vertical",
                "width": "32px",
                "height": "32px",
                "cornerRadius": "16px",
                "backgroundColor": "#FFFFFF",
                "alignItems": "center",
                "justifyContent": "center",
                "contents": [
                    {
                        "type": "text",
                        "text": ">",
                        "weight": "bold",
                        "size": "lg",
                        "color": accent,
                        "align": "center",
                    }
                ],
            },
        ],
    }


def hr_menu_message() -> dict[str, Any]:
    buttons = [
        hr_menu_button("1", "เธ•เธฒเธฃเธฒเธเธเธฒเธ", "เธ”เธนเธ•เธฒเธฃเธฒเธเธเธฒเธเธเธญเธเธเธธเธ“", "เธ•เธฒเธฃเธฒเธเธเธฒเธ", "#F3ECFF", "#8B5CF6"),
        hr_menu_button("2", "เธฅเธฒเธเนเธงเธข", "เนเธเนเธเธฅเธฒเธเนเธงเธข / เธเธฑเธเธ—เธถเธเธเธฒเธฃเธฅเธฒเธเนเธงเธข", "เธฅเธฒเธเนเธงเธข", "#E8FBF7", "#2DD4BF"),
        hr_menu_button("3", "เธฅเธฒเธเธดเธ", "เนเธเนเธเธฅเธฒเธเธดเธ / เธเธฑเธเธ—เธถเธเธเธฒเธฃเธฅเธฒเธเธดเธ", "เธฅเธฒเธเธดเธ", "#FFF7E6", "#F59E0B"),
        hr_menu_button("4", "เนเธเนเธเธเธญเธงเธฑเธเธซเธขเธธเธ”เธฅเนเธงเธเธซเธเนเธฒ", "เธเธญเธงเธฑเธเธซเธขเธธเธ”เธฅเนเธงเธเธซเธเนเธฒ / เธงเธฒเธเนเธเธเธงเธฑเธเธซเธขเธธเธ”", "เนเธเนเธเธเธญเธงเธฑเธเธซเธขเธธเธ”เธฅเนเธงเธเธซเธเนเธฒ", "#FFF0F6", "#EC4899"),
        hr_menu_button("5", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเน€เธงเธฅเธฒเน€เธเนเธฒ-เธญเธญเธเธเธฒเธ", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเน€เธงเธฅเธฒเน€เธเนเธฒ-เธญเธญเธเธเธฒเธเธฅเนเธงเธเธซเธเนเธฒ", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเน€เธงเธฅเธฒเน€เธเนเธฒ-เธญเธญเธเธเธฒเธ", "#EFF6FF", "#3B82F6"),
        hr_menu_button("6", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเธงเธฑเธเธ—เธณเธเธฒเธ", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเธงเธฑเธเธ—เธณเธเธฒเธ / เธชเธฅเธฑเธเธงเธฑเธเธ—เธณเธเธฒเธ", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเธงเธฑเธเธ—เธณเธเธฒเธ", "#F5F3FF", "#A855F7"),
    ]
    return {
        "type": "flex",
        "altText": "เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเน€เธกเธเธน HR",
        "contents": {
            "type": "bubble",
            "size": "giga",
            "body": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#FFFBF7",
                "paddingAll": "16px",
                "spacing": "10px",
                "contents": [
                    {
                        "type": "text",
                        "text": "เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเน€เธกเธเธน HR",
                        "weight": "bold",
                        "size": "xl",
                        "color": "#111827",
                    },
                    {
                        "type": "text",
                        "text": "เนเธ•เธฐเธเธธเนเธกเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเนเธเนเธเธฒเธเนเธ”เนเน€เธฅเธขเธเนเธฐ",
                        "size": "sm",
                        "color": "#64748B",
                        "margin": "xs",
                    },
                    {"type": "separator", "margin": "md"},
                    *buttons,
                ],
            },
        },
    }



def hr_request_blank(request_type: str, line_user_id: str = "") -> dict[str, Any]:
    today = dt.date.today().isoformat()
    return {
        "request_type": request_type,
        "employee_name": get_line_display_name(line_user_id),
        "start_date": today,
        "end_date": today,
        "work_date": today,
        "old_date": "",
        "new_date": "",
        "old_time": "",
        "new_time": "",
        "reason": "",
        "note": "",
        "status": "เธฃเธญเธญเธเธธเธกเธฑเธ•เธด",
    }


def hr_request_form(data: dict[str, Any]) -> str:
    request_type = str(data.get("request_type") or "เธเธณเธเธญ HR")
    common = (
        f"==== {request_type} ====\n"
        "เธเธฃเธธเธ“เธฒเธเธฃเธญเธ/เนเธเนเนเธเน€เธเธเธฒเธฐเธซเธฑเธงเธเนเธญเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃ เนเธฅเนเธงเธชเนเธเธเธฅเธฑเธเธกเธฒเนเธ”เนเน€เธฅเธขเธเนเธฐ\n\n"
        f"เธเธทเนเธญเธเธเธฑเธเธเธฒเธ: {data.get('employee_name') or ''}\n"
    )
    if request_type in {"เธฅเธฒเธเนเธงเธข", "เธฅเธฒเธเธดเธ", "เนเธเนเธเธเธญเธงเธฑเธเธซเธขเธธเธ”เธฅเนเธงเธเธซเธเนเธฒ"}:
        return (
            common +
            f"เธงเธฑเธเธ—เธตเนเน€เธฃเธดเนเธก: {data.get('start_date') or ''}\n"
            f"เธงเธฑเธเธ—เธตเนเธชเธดเนเธเธชเธธเธ”: {data.get('end_date') or ''}\n"
            f"เน€เธซเธ•เธธเธเธฅ: {data.get('reason') or ''}\n"
            f"เธซเธกเธฒเธขเน€เธซเธ•เธธ: {data.get('note') or ''}"
        )
    if request_type == "เนเธเนเธเน€เธเธฅเธตเนเธขเธเน€เธงเธฅเธฒเน€เธเนเธฒ-เธญเธญเธเธเธฒเธ":
        return (
            common +
            f"เธงเธฑเธเธ—เธตเนเธ—เธณเธเธฒเธ: {data.get('work_date') or ''}\n"
            f"เน€เธงเธฅเธฒเน€เธ”เธดเธก: {data.get('old_time') or ''}\n"
            f"เน€เธงเธฅเธฒเนเธซเธกเน: {data.get('new_time') or ''}\n"
            f"เน€เธซเธ•เธธเธเธฅ: {data.get('reason') or ''}\n"
            f"เธซเธกเธฒเธขเน€เธซเธ•เธธ: {data.get('note') or ''}"
        )
    if request_type == "เนเธเนเธเน€เธเธฅเธตเนเธขเธเธงเธฑเธเธ—เธณเธเธฒเธ":
        return (
            common +
            f"เธงเธฑเธเธ—เธตเนเน€เธ”เธดเธก: {data.get('old_date') or ''}\n"
            f"เธงเธฑเธเธ—เธตเนเนเธซเธกเน: {data.get('new_date') or ''}\n"
            f"เน€เธซเธ•เธธเธเธฅ: {data.get('reason') or ''}\n"
            f"เธซเธกเธฒเธขเน€เธซเธ•เธธ: {data.get('note') or ''}"
        )
    return common + f"เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”: {data.get('note') or ''}"


def format_hr_request(data: dict[str, Any]) -> str:
    return (
        f"==== เธเธณเธเธญ HR ====\n"
        f"เธเธฃเธฐเน€เธ เธ—: {data.get('request_type') or '-'}\n"
        f"เธเธทเนเธญเธเธเธฑเธเธเธฒเธ: {data.get('employee_name') or '-'}\n"
        f"เธงเธฑเธเธ—เธตเนเน€เธฃเธดเนเธก: {data.get('start_date') or '-'}\n"
        f"เธงเธฑเธเธ—เธตเนเธชเธดเนเธเธชเธธเธ”: {data.get('end_date') or '-'}\n"
        f"เธงเธฑเธเธ—เธตเนเธ—เธณเธเธฒเธ: {data.get('work_date') or '-'}\n"
        f"เธงเธฑเธเธ—เธตเนเน€เธ”เธดเธก: {data.get('old_date') or '-'}\n"
        f"เธงเธฑเธเธ—เธตเนเนเธซเธกเน: {data.get('new_date') or '-'}\n"
        f"เน€เธงเธฅเธฒเน€เธ”เธดเธก: {data.get('old_time') or '-'}\n"
        f"เน€เธงเธฅเธฒเนเธซเธกเน: {data.get('new_time') or '-'}\n"
        f"เน€เธซเธ•เธธเธเธฅ: {data.get('reason') or '-'}\n"
        f"เธซเธกเธฒเธขเน€เธซเธ•เธธ: {data.get('note') or '-'}\n"
        f"เธชเธ–เธฒเธเธฐ: {data.get('status') or 'เธฃเธญเธญเธเธธเธกเธฑเธ•เธด'}"
    )


def hr_confirm_message(data: dict[str, Any]) -> dict[str, Any]:
    return quick_reply_text_message(
        format_hr_request(data) + "\n\n"
        "เธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเธชเธญเธเธเนเธญเธกเธนเธฅเธเนเธญเธเธชเนเธเธเธณเธเธญเธญเธเธธเธกเธฑเธ•เธด\n"
        "1 = เธขเธทเธเธขเธฑเธเธชเนเธเธเธณเธเธญ\n"
        "2 = เนเธเนเนเธเธเนเธญเธกเธนเธฅ",
        [
            ("1 เธขเธทเธเธขเธฑเธ", "1"),
            ("2 เนเธเนเนเธ", "2"),
        ],
    )


def parse_hr_request_text(text: str, data: dict[str, Any]) -> dict[str, Any]:
    updated = dict(data)
    aliases = {
        "เธเธฃเธฐเน€เธ เธ—": "request_type",
        "เธเธทเนเธญเธเธเธฑเธเธเธฒเธ": "employee_name",
        "เธเธทเนเธญ": "employee_name",
        "เธงเธฑเธเธ—เธตเนเน€เธฃเธดเนเธก": "start_date",
        "เธงเธฑเธเธ—เธตเนเธชเธดเนเธเธชเธธเธ”": "end_date",
        "เธงเธฑเธเธ—เธตเนเธฅเธฒ": "start_date",
        "เธงเธฑเธเธ—เธตเนเธ—เธณเธเธฒเธ": "work_date",
        "เธงเธฑเธเธ—เธตเนเน€เธ”เธดเธก": "old_date",
        "เธงเธฑเธเธ—เธตเนเนเธซเธกเน": "new_date",
        "เน€เธงเธฅเธฒเน€เธ”เธดเธก": "old_time",
        "เน€เธงเธฅเธฒเนเธซเธกเน": "new_time",
        "เน€เธซเธ•เธธเธเธฅ": "reason",
        "เธซเธกเธฒเธขเน€เธซเธ•เธธ": "note",
        "เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”": "note",
    }
    normalized_aliases = {re.sub(r"[\s/_-]+", "", k.lower()): v for k, v in aliases.items()}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^(เนเธเนเนเธ|เน€เธเธฅเธตเนเธขเธ)\s*", "", line)
        match = re.match(r"^([^:=:\-]+)\s*[:=:\-]\s*(.*)$", line)
        if not match:
            match = re.match(r"^(\S+)\s+(.+)$", line)
        if not match:
            continue
        raw_key = match.group(1).strip().lower()
        value = match.group(2).strip()
        key = aliases.get(raw_key) or normalized_aliases.get(re.sub(r"[\s/_-]+", "", raw_key))
        if not key:
            continue
        if key.endswith("date") or key in {"start_date", "end_date", "work_date", "old_date", "new_date"}:
            parsed = parse_date(value)
            updated[key] = parsed.isoformat() if parsed else value
        else:
            updated[key] = value
    return updated


def abort_flow_message(reason: str) -> str:
    return (
        f"{reason}\n\n"
        "เธฃเธฐเธเธเธซเธขเธธเธ”เธเธฒเธเธฃเธฒเธขเธเธฒเธฃเธเธตเนเนเธซเนเนเธฅเนเธงเธเนเธฐ เธชเธฒเธกเธฒเธฃเธ–เน€เธฃเธดเนเธกเธ—เธณเธฃเธฒเธขเธเธฒเธฃเนเธซเธกเนเนเธ”เนเน€เธฅเธข\n\n"
        + menu_text()
    )


def blank_manual_entry(transaction_type: str) -> dict[str, Any]:
    normalized_type = "Revenue" if str(transaction_type).lower() == "revenue" else "Expense"
    category = CONFIG.get("default_revenue_category", "Sales") if normalized_type == "Revenue" else CONFIG.get("default_category", "Other")
    return {
        "transaction_type": normalized_type,
        "document_type": "",
        "date": dt.date.today(),
        "invoice_no": "-",
        "vendor": "",
        "description": "Manual entry from LINE",
        "category": category,
        "before_vat": 0,
        "vat": 0,
        "withholding_tax": 0,
        "total": 0,
        "claimable": CONFIG.get("default_claimable", "Yes"),
        "confidence": 0,
        "raw_text": "",
        "submitter_name": "",
    }


def manual_entry_form(data: dict[str, Any]) -> str:
    return (
        "OCR เธญเนเธฒเธเน€เธญเธเธชเธฒเธฃเนเธกเนเธชเธณเน€เธฃเนเธเธซเธฃเธทเธญเนเธเนเน€เธงเธฅเธฒเธเธฒเธเน€เธเธดเธเนเธเธเนเธฐ\n"
        "เธเธฃเธธเธ“เธฒเธเธฃเธญเธเธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธ•เธฒเธกเนเธเธเธเธญเธฃเนเธกเธเธตเน เนเธฅเนเธงเธชเนเธเธเธฅเธฑเธเธกเธฒเนเธ”เนเน€เธฅเธขเธเนเธฐ\n\n"
        f"เธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ: {data.get('document_type') or ''}\n"
        f"เธงเธฑเธเธ—เธตเน: {data.get('date') or ''}\n"
        f"เน€เธฅเธเธ—เธตเนเธเธดเธฅ: {normalize_invoice_no(data.get('invoice_no'))}\n"
        f"เธเธทเนเธญเธฃเนเธฒเธ/เธเธนเนเธเนเธฒ: {data.get('vendor') or ''}\n"
        f"เธเธนเนเธเธณเธชเนเธเน€เธญเธเธชเธฒเธฃ: {data.get('submitter_name') or ''}\n"
        f"เธซเธกเธงเธ”: {data.get('category') or ''}\n"
        f"เธขเธญเธ”เธเนเธญเธ VAT: {float(data.get('before_vat') or 0):.2f}\n"
        f"VAT: {float(data.get('vat') or 0):.2f}\n"
        f"เธ เธฒเธฉเธตเธซเธฑเธ เธ“ เธ—เธตเนเธเนเธฒเธข: {float(data.get('withholding_tax') or 0):.2f}\n"
        f"เธขเธญเธ”เธฃเธงเธก: {float(data.get('total') or 0):.2f}"
    )


def confirm_edit_button(label: str, text: str, background: str, icon: str) -> dict[str, Any]:
    return {
        "type": "box",
        "layout": "horizontal",
        "cornerRadius": "24px",
        "backgroundColor": background,
        "paddingAll": "18px",
        "spacing": "14px",
        "action": {"type": "message", "label": label[:20], "text": text},
        "contents": [
            {
                "type": "box",
                "layout": "vertical",
                "width": "52px",
                "height": "52px",
                "cornerRadius": "26px",
                "backgroundColor": "#FFFFFF",
                "alignItems": "center",
                "justifyContent": "center",
                "contents": [
                    {
                        "type": "text",
                        "text": icon,
                        "weight": "bold",
                        "size": "xxl",
                        "color": background,
                        "align": "center",
                    }
                ],
            },
            {
                "type": "text",
                "text": label,
                "weight": "bold",
                "size": "xxl",
                "color": "#FFFFFF",
                "gravity": "center",
                "wrap": True,
                "flex": 1,
            },
        ],
    }


def confirm_edit_buttons_message() -> dict[str, Any]:
    return {
        "type": "flex",
        "altText": "ยืนยันหรือแก้ไขข้อมูล",
        "contents": {
            "type": "bubble",
            "size": "mega",
            "body": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#FFFFFF",
                "paddingAll": "18px",
                "spacing": "16px",
                "contents": [
                    confirm_edit_button("1. ยืนยัน", "1", "#10B981", "OK"),
                    confirm_edit_button("2. แก้ไข", "2", "#8B5CF6", "Edit"),
                ],
            },
        },
    }


def approval_buttons_message(request_id: str) -> dict[str, Any]:
    return {
        "type": "flex",
        "altText": "อนุมัติหรือไม่อนุมัติ",
        "contents": {
            "type": "bubble",
            "size": "mega",
            "body": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#FFFFFF",
                "paddingAll": "18px",
                "spacing": "16px",
                "contents": [
                    confirm_edit_button("1. อนุมัติ", f"HR_APPROVE:{request_id}", "#10B981", "OK"),
                    confirm_edit_button("2. ไม่อนุมัติ", f"HR_REJECT:{request_id}", "#EF4444", "No"),
                ],
            },
        },
    }


def confirmation_prompt(data: dict[str, Any]) -> list[dict[str, Any]]:
    detail_text = (
        format_parsed_details(data, "เธเธดเธฅเธเธณเน€เธเนเธฒ") + "\n\n"
        "เธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเธชเธญเธเธเนเธญเธกเธนเธฅเธเนเธญเธเธเธฑเธเธ—เธถเธเธฅเธ Google Sheet เธเนเธฐ\n"
        "เธเธ”เธเธธเนเธกเธ”เนเธฒเธเธฅเนเธฒเธ เธซเธฃเธทเธญเธเธดเธกเธเนเน€เธฅเธ 1/2 เนเธ”เนเน€เธฅเธข"
    )
    return [text_message(detail_text), confirm_edit_buttons_message()]


def confirm_pending_to_google(line_user_id: str, state: dict[str, Any], public_base_url: str) -> str | list[dict[str, Any]]:
    pending = deserialize_data(state["pending_data"])
    if not str(pending.get("submitter_name") or "").strip():
        state["mode"] = "awaiting_submitter_name"
        state["pending_data"] = serialize_data(pending)
        set_user_state(line_user_id, state)
        return "เธเธฃเธธเธ“เธฒเธฃเธฐเธเธธเธเธทเนเธญเธเธนเนเธเธณเธชเนเธเน€เธญเธเธชเธฒเธฃเธเนเธญเธเธเธฑเธเธ—เธถเธเธเนเธฐ"
    if not state.get("duplicate_checked"):
        try:
            matches = search_google_sheet_by_total(pending.get("total"))
        except Exception as exc:
            runtime_log(f"Duplicate total check failed: {exc}")
            matches = []
        if matches:
            state["mode"] = "awaiting_duplicate_confirmation"
            state["duplicate_matches"] = matches[:10]
            state["pending_data"] = serialize_data(pending)
            set_user_state(line_user_id, state)
            return (
                format_google_sheet_matches(matches, "เธเธเธฃเธฒเธขเธเธฒเธฃเธขเธญเธ”เธฃเธงเธกเธเนเธณเนเธ Google Sheet") +
                "\n\nเธฃเธฒเธขเธเธฒเธฃเธเธตเนเน€เธเนเธเธเนเธญเธกเธนเธฅเธ•เธฑเธงเน€เธ”เธตเธขเธงเธเธฑเธเธซเธฃเธทเธญเนเธกเน?\n"
                "เธ•เธญเธ 1 = เนเธเน เน€เธเนเธเธฃเธฒเธขเธเธฒเธฃเน€เธ”เธตเธขเธงเธเธฑเธ\n"
                "เธ•เธญเธ 2 = เนเธกเนเนเธเน เธเธฑเธเธ—เธถเธเน€เธเนเธเธฃเธฒเธขเธเธฒเธฃเนเธซเธกเน"
            )
    image_path = Path(state.get("image_path", ""))
    if not image_path.exists():
        return "เนเธกเนเธเธเนเธเธฅเนเธฃเธนเธเน€เธญเธเธชเธฒเธฃเน€เธ”เธดเธกเธเนเธฐ เธเธฃเธธเธ“เธฒเธชเนเธเน€เธญเธเธชเธฒเธฃเนเธซเธกเนเธญเธตเธเธเธฃเธฑเนเธ"
    result = None
    try:
        result = send_to_google_sheet(pending, image_path, line_user_id, public_base_url)
    except Exception as exc:
        runtime_log(f"Google Sheet save failed: {exc}")
        clear_user_state(line_user_id)
        return abort_flow_message(f"Google Sheet: เธขเธฑเธเนเธกเนเธชเธณเน€เธฃเนเธ ({exc})")
    notify_accounting_import_approver(pending, line_user_id)
    summary_image = render_row_summary_image("Google Sheet", "-", pending, "เธเธดเธฅเธเธณเน€เธเนเธฒ")
    messages = [
        text_message("เธเธฑเธเธ—เธถเธเน€เธฃเธตเธขเธเธฃเนเธญเธข\nGoogle Sheet: เธเธฑเธเธ—เธถเธเธชเธณเน€เธฃเนเธ"),
        image_message(public_file_url(public_base_url, summary_image)),
    ]
    substitute_match_data = substitute_match_from_pending(pending, result)
    if can_create_substitute_receipt(substitute_match_data):
        set_user_state(
            line_user_id,
            {
                "mode": "awaiting_substitute_receipt_decision",
                "substitute_match": substitute_match_data,
            },
        )
        messages.append(
            quick_reply_text_message(
                "เธเธฑเธเธ—เธถเธเน€เธเนเธเธเธดเธฅ/เธเธดเธฅเน€เธเธดเธเธชเธ”เน€เธฃเธตเธขเธเธฃเนเธญเธขเธเนเธฐ\n\n"
                "เธ•เนเธญเธเธเธฒเธฃเธชเธฃเนเธฒเธเนเธเนเธ—เธเธชเธณเธซเธฃเธฑเธเธเธดเธกเธเนเน€เธเนเธเน€เธเนเธ hard copy เนเธซเธกเธเธฐ?\n"
                "เน€เธฅเธทเธญเธเธเธธเนเธกเธ”เนเธฒเธเธฅเนเธฒเธ เธซเธฃเธทเธญเธเธดเธกเธเนเน€เธฅเธเธ•เธญเธเธเธฅเธฑเธเนเธ”เนเน€เธฅเธข\n\n"
                "1 = เธ•เนเธญเธเธเธฒเธฃเธชเธฃเนเธฒเธเนเธเนเธ—เธ\n"
                "2 = เนเธกเนเธ•เนเธญเธเธเธฒเธฃ",
                [
                    ("๐งพ 1 เธชเธฃเนเธฒเธเนเธเนเธ—เธ", "1"),
                    ("เนเธกเนเธชเธฃเนเธฒเธเนเธเนเธ—เธ", "2"),
                ],
            )
        )
    else:
        clear_user_state(line_user_id)
    runtime_log(
        f"Confirmed LINE receipt -> Google Sheet "
        f"type={pending.get('transaction_type')} date={pending['date']} total={pending['total']}"
    )
    return messages


def parse_correction_text(text: str, data: dict[str, Any]) -> dict[str, Any]:
    updated = dict(data)
    aliases = {
        "date": "date",
        "เธงเธฑเธเธ—เธตเน": "date",
        "invoice": "invoice_no",
        "invoice no": "invoice_no",
        "invoice number": "invoice_no",
        "เน€เธฅเธเธ—เธตเนเน€เธญเธเธชเธฒเธฃ": "invoice_no",
        "เน€เธฅเธเธ—เธตเนเธเธดเธฅ": "invoice_no",
        "เน€เธฅเธเธ—เธตเนเนเธเน€เธชเธฃเนเธ": "invoice_no",
        "เธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ": "document_type",
        "เธเธฃเธฐเน€เธ เธ—เธเธดเธฅ": "document_type",
        "เธเธเธดเธ”เน€เธญเธเธชเธฒเธฃ": "document_type",
        "document type": "document_type",
        "vendor": "vendor",
        "supplier": "vendor",
        "เธฃเนเธฒเธ": "vendor",
        "เธเธนเนเธเธฒเธข": "vendor",
        "เธเธนเนเธเนเธฒ": "vendor",
        "category": "category",
        "เธซเธกเธงเธ”": "category",
        "description": "description",
        "เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”": "description",
        "before vat": "before_vat",
        "เธขเธญเธ”เธเนเธญเธ vat": "before_vat",
        "เธขเธญเธ”เธเนเธญเธเธ เธฒเธฉเธต": "before_vat",
        "vat": "vat",
        "เธ เธฒเธฉเธต": "vat",
        "total": "total",
        "เธขเธญเธ”เธฃเธงเธก": "total",
    }
    changed = False
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match(r"^([^:=๏ผ]+)\s*[:=๏ผ]\s*(.+)$", line)
        if not match:
            continue
        raw_key, value = match.group(1).strip().lower(), match.group(2).strip()
        key = aliases.get(raw_key)
        if not key:
            continue
        if key == "date":
            parsed_date = parse_date(value)
            if parsed_date:
                updated[key] = parsed_date
                changed = True
        elif key in {"before_vat", "vat", "total"}:
            amount = normalize_amount(value)
            if amount is not None:
                updated[key] = round(amount, 2)
                changed = True
        else:
            updated[key] = value
            changed = True

    if changed:
        if updated.get("total") and updated.get("vat") and not updated.get("before_vat"):
            updated["before_vat"] = round(float(updated["total"]) - float(updated["vat"]), 2)
        elif updated.get("before_vat") and updated.get("vat"):
            updated["total"] = round(float(updated["before_vat"]) + float(updated["vat"]), 2)
    return updated


def parse_correction_text_v2(text: str, data: dict[str, Any]) -> dict[str, Any]:
    updated = dict(data)
    changed_fields: set[str] = set()

    aliases = {
        "date": "date",
        "\u0e27\u0e31\u0e19\u0e17\u0e35\u0e48": "date",
        "invoice": "invoice_no",
        "invoice no": "invoice_no",
        "invoice number": "invoice_no",
        "\u0e40\u0e25\u0e02\u0e17\u0e35\u0e48\u0e40\u0e2d\u0e01\u0e2a\u0e32\u0e23": "invoice_no",
        "\u0e40\u0e25\u0e02\u0e17\u0e35\u0e48\u0e1a\u0e34\u0e25": "invoice_no",
        "\u0e40\u0e25\u0e02\u0e17\u0e35\u0e48\u0e43\u0e1a\u0e40\u0e2a\u0e23\u0e47\u0e08": "invoice_no",
        "type": "transaction_type",
        "\u0e1b\u0e23\u0e30\u0e40\u0e20\u0e17": "transaction_type",
        "\u0e1b\u0e23\u0e30\u0e40\u0e20\u0e17\u0e23\u0e32\u0e22\u0e01\u0e32\u0e23": "transaction_type",
        "\u0e1b\u0e23\u0e30\u0e40\u0e20\u0e17\u0e40\u0e2d\u0e01\u0e2a\u0e32\u0e23": "document_type",
        "\u0e1b\u0e23\u0e30\u0e40\u0e20\u0e17\u0e1a\u0e34\u0e25": "document_type",
        "\u0e0a\u0e19\u0e34\u0e14\u0e40\u0e2d\u0e01\u0e2a\u0e32\u0e23": "document_type",
        "document type": "document_type",
        "vendor": "vendor",
        "supplier": "vendor",
        "\u0e23\u0e49\u0e32\u0e19": "vendor",
        "\u0e1c\u0e39\u0e49\u0e02\u0e32\u0e22": "vendor",
        "\u0e04\u0e39\u0e48\u0e04\u0e49\u0e32": "vendor",
        "\u0e0a\u0e37\u0e48\u0e2d\u0e23\u0e49\u0e32\u0e19": "vendor",
        "\u0e0a\u0e37\u0e48\u0e2d\u0e23\u0e49\u0e32\u0e19/\u0e04\u0e39\u0e48\u0e04\u0e49\u0e32": "vendor",
        "\u0e0a\u0e37\u0e48\u0e2d\u0e1c\u0e39\u0e49\u0e02\u0e32\u0e22": "vendor",
        "\u0e1c\u0e39\u0e49\u0e19\u0e33\u0e2a\u0e48\u0e07\u0e40\u0e2d\u0e01\u0e2a\u0e32\u0e23": "submitter_name",
        "\u0e0a\u0e37\u0e48\u0e2d\u0e1c\u0e39\u0e49\u0e19\u0e33\u0e2a\u0e48\u0e07": "submitter_name",
        "submitter": "submitter_name",
        "submitter name": "submitter_name",
        "category": "category",
        "\u0e2b\u0e21\u0e27\u0e14": "category",
        "description": "description",
        "\u0e23\u0e32\u0e22\u0e25\u0e30\u0e40\u0e2d\u0e35\u0e22\u0e14": "description",
        "before vat": "before_vat",
        "\u0e22\u0e2d\u0e14\u0e01\u0e48\u0e2d\u0e19 vat": "before_vat",
        "\u0e22\u0e2d\u0e14\u0e01\u0e48\u0e2d\u0e19\u0e20\u0e32\u0e29\u0e35": "before_vat",
        "vat": "vat",
        "\u0e20\u0e32\u0e29\u0e35": "vat",
        "withholding tax": "withholding_tax",
        "wht": "withholding_tax",
        "tax withheld": "withholding_tax",
        "\u0e20\u0e32\u0e29\u0e35\u0e2b\u0e31\u0e01 \u0e13 \u0e17\u0e35\u0e48\u0e08\u0e48\u0e32\u0e22": "withholding_tax",
        "\u0e20\u0e32\u0e29\u0e35\u0e2b\u0e31\u0e01\u0e13\u0e17\u0e35\u0e48\u0e08\u0e48\u0e32\u0e22": "withholding_tax",
        "total": "total",
        "\u0e22\u0e2d\u0e14\u0e23\u0e27\u0e21": "total",
    }
    normalized_aliases = {
        re.sub(r"[\s/_-]+", "", key.lower()): value
        for key, value in aliases.items()
    }
    keyword_pattern = "|".join(re.escape(key) for key in sorted(aliases, key=len, reverse=True))

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^[\-\*\u2022]\s*", "", line)
        line = re.sub(r"^(เนเธเนเนเธ|เน€เธเธฅเธตเนเธขเธ)\s*", "", line)
        match = re.match(r"^([^:=:\-]+)\s*[:=:\-]\s*(.+)$", line)
        if not match:
            match = re.match(rf"^({keyword_pattern})\s+(.+)$", line, flags=re.IGNORECASE)
        if not match:
            continue

        raw_key = match.group(1).strip().lower()
        value = match.group(2).strip()
        normalized_key = re.sub(r"[\s/_-]+", "", raw_key)
        key = aliases.get(raw_key) or normalized_aliases.get(normalized_key)
        if not key:
            continue

        if key == "date":
            parsed_date = parse_date(value)
            if not parsed_date:
                continue
            updated[key] = parsed_date
        elif key in {"before_vat", "vat", "withholding_tax", "total"}:
            amount = normalize_amount(value)
            if amount is None:
                continue
            updated[key] = round(amount, 2)
        elif key == "document_type":
            normalized_doc_type = value.strip()
            if normalized_doc_type in {"bill"}:
                normalized_doc_type = "\u0e1a\u0e34\u0e25"
            elif normalized_doc_type in {"receipt"}:
                normalized_doc_type = "\u0e43\u0e1a\u0e40\u0e2a\u0e23\u0e47\u0e08"
            elif normalized_doc_type in {"tax invoice", "invoice"}:
                normalized_doc_type = "\u0e43\u0e1a\u0e01\u0e33\u0e01\u0e31\u0e1a\u0e20\u0e32\u0e29\u0e35"
            updated[key] = normalized_doc_type
        elif key == "transaction_type":
            normalized_type = value.strip().lower()
            if normalized_type in {"revenue", "\u0e23\u0e32\u0e22\u0e23\u0e31\u0e1a", "\u0e1a\u0e34\u0e25\u0e23\u0e32\u0e22\u0e23\u0e31\u0e1a"}:
                updated[key] = "Revenue"
            elif normalized_type in {"expense", "expenses", "\u0e23\u0e32\u0e22\u0e08\u0e48\u0e32\u0e22", "\u0e1a\u0e34\u0e25\u0e23\u0e32\u0e22\u0e08\u0e48\u0e32\u0e22"}:
                updated[key] = "Expense"
            else:
                updated[key] = value
        elif key == "invoice_no":
            updated[key] = normalize_invoice_no(value)
        else:
            updated[key] = value
        changed_fields.add(key)

    if changed_fields:
        before_changed = "before_vat" in changed_fields
        vat_changed = "vat" in changed_fields
        total_changed = "total" in changed_fields
        before = float(updated.get("before_vat") or 0)
        vat = float(updated.get("vat") or 0)
        total = float(updated.get("total") or 0)
        if before_changed and not vat_changed and not total_changed:
            vat = round(before * float(CONFIG.get("vat_rate", 0.07)), 2)
            total = round(before + vat, 2)
            updated["vat"] = vat
            updated["total"] = total
        elif before_changed and vat_changed and not total_changed:
            updated["total"] = round(before + vat, 2)
        elif total_changed and vat_changed and not before_changed:
            updated["before_vat"] = round(total - vat, 2)
        elif total_changed and before_changed and not vat_changed:
            updated["vat"] = round(total - before, 2)
        elif not total_changed and (before_changed or vat_changed):
            updated["total"] = round(before + vat, 2)

    return updated


def find_font(size: int):
    from PIL import ImageFont

    candidates = [
        r"C:\Windows\Fonts\tahoma.ttf",
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
        "/usr/share/fonts/truetype/tlwg/Garuda.ttf",
        "/usr/share/fonts/truetype/tlwg/Garuda-Bold.ttf",
        "/usr/share/fonts/truetype/tlwg/Loma.ttf",
        "/usr/share/fonts/truetype/tlwg/Loma-Bold.ttf",
        "/usr/share/fonts/truetype/tlwg/TlwgTypist.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansThaiLooped-Regular.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansThaiLoopedUI-Regular.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansThai-Regular.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansThaiUI-Regular.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansThai-Regular.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansThaiUI-Regular.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    for root in (Path("/usr/share/fonts/truetype/noto"), Path("/usr/share/fonts/opentype/noto")):
        if root.exists():
            for pattern in ("NotoSansThai*.ttf", "NotoSansThai*.otf", "NotoSerifThai*.ttf", "NotoSerifThai*.otf"):
                for candidate in sorted(root.glob(pattern)):
                    return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def wrap_text(text: str, max_chars: int) -> list[str]:
    text = str(text or "-")
    if len(text) <= max_chars:
        return [text]
    lines: list[str] = []
    current = ""
    for part in text.split():
        if len(current) + len(part) + 1 <= max_chars:
            current = f"{current} {part}".strip()
        else:
            if current:
                lines.append(current)
            current = part
    if current:
        lines.append(current)
    if not lines:
        lines = [text[i:i + max_chars] for i in range(0, len(text), max_chars)]
    return lines[:3]


def render_row_summary_image(sheet_name: str, row: int, data: dict[str, Any], heading: str) -> Path:
    from PIL import Image, ImageDraw

    out_dir = reply_image_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"excel_row_{row}_{uuid.uuid4().hex[:8]}.png"
    width, height = 1280, 980
    image = Image.new("RGB", (width, height), "#F8FAFC")
    draw = ImageDraw.Draw(image)
    title_font = find_font(36)
    label_font = find_font(25)
    value_font = find_font(25)
    small_font = find_font(21)

    accent = "#BBF7D0" if data.get("transaction_type") == "Revenue" else "#FECACA"
    draw.rounded_rectangle((28, 28, width - 28, 120), radius=22, fill=accent)
    draw.text((58, 52), heading, fill="#111827", font=title_font)
    draw.text((width - 440, 62), f"{sheet_name}!Row {row}", fill="#374151", font=small_font)

    fields = [
        ("เธเธฃเธฐเน€เธ เธ—", data.get("transaction_type", "-")),
        ("เธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ", data.get("document_type") or "-"),
        ("เธงเธฑเธเธ—เธตเน", str(data.get("date", "-"))),
        ("เน€เธฅเธเธ—เธตเนเธเธดเธฅ", data.get("invoice_no") or "-"),
        ("เธเธทเนเธญเธฃเนเธฒเธ/เธเธนเนเธเนเธฒ", data.get("vendor") or "-"),
        ("เธเธนเนเธเธณเธชเนเธเน€เธญเธเธชเธฒเธฃ", data.get("submitter_name") or "-"),
        ("เธซเธกเธงเธ”", data.get("category") or "-"),
        ("เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”", data.get("description") or "-"),
        ("เธขเธญเธ”เธเนเธญเธ VAT", f"{float(data.get('before_vat') or 0):,.2f}"),
        ("VAT", f"{float(data.get('vat') or 0):,.2f}"),
        ("เธ เธฒเธฉเธตเธซเธฑเธ เธ“ เธ—เธตเนเธเนเธฒเธข", f"{float(data.get('withholding_tax') or 0):,.2f}"),
        ("เธขเธญเธ”เธฃเธงเธก", f"{float(data.get('total') or 0):,.2f}"),
    ]
    y = 155
    for label, value in fields:
        wrapped_value = wrap_text(str(value), 54)
        box_height = max(58, 30 + (len(wrapped_value) * 28))
        draw.rounded_rectangle((42, y - 8, width - 42, y + box_height), radius=10, fill="#FFFFFF", outline="#E5E7EB")
        draw.text((70, y), label, fill="#475569", font=label_font)
        line_y = y
        for value_line in wrapped_value:
            draw.text((395, line_y), value_line, fill="#111827", font=value_font)
            line_y += 28
        y += box_height + 8
    image.save(path, quality=92)
    return path


def is_substitute_receipt_doc(document_type: Any) -> bool:
    normalized = re.sub(r"\s+", "", str(document_type or "").lower())
    return (
        normalized in {"เธเธดเธฅ", "เธเธดเธฅเน€เธเธดเธเธชเธ”", "bill", "cashbill"}
        or "เธเธดเธฅ" in normalized
    )


def can_create_substitute_receipt(item: dict[str, Any]) -> bool:
    item_type = str(item.get("type") or "").lower()
    return item_type == "expense" and (
        is_substitute_receipt_doc(item.get("documentType"))
        or not str(item.get("documentType") or "").strip()
    )


def substitute_match_from_pending(data: dict[str, Any], result: dict[str, Any] | None = None) -> dict[str, Any]:
    result = result or {}
    return {
        "row": result.get("row") or "-",
        "sheetName": result.get("sheetName") or result.get("sheet") or "Google Sheet",
        "date": data.get("date") or "-",
        "type": data.get("transaction_type") or data.get("type") or "Expense",
        "invoiceNo": data.get("invoice_no") or data.get("invoiceNo") or "-",
        "vendor": data.get("vendor") or "-",
        "description": data.get("description") or "-",
        "category": data.get("category") or "-",
        "beforeVat": data.get("before_vat") or data.get("beforeVat") or 0,
        "vat": data.get("vat") or 0,
        "withholdingTax": data.get("withholding_tax") or data.get("withholdingTax") or 0,
        "total": data.get("total") or 0,
        "documentType": data.get("document_type") or data.get("documentType") or "-",
        "submitterName": data.get("submitter_name") or data.get("submitterName") or "-",
    }


def substitute_data_from_match(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "date": item.get("date") or "-",
        "transaction_type": item.get("type") or "Expense",
        "invoice_no": item.get("invoiceNo") or "-",
        "vendor": item.get("vendor") or "-",
        "description": item.get("description") or "เธเนเธฒเนเธเนเธเนเธฒเธขเธ•เธฒเธกเธเธดเธฅ/เธเธดเธฅเน€เธเธดเธเธชเธ”",
        "category": item.get("category") or "-",
        "before_vat": float(item.get("beforeVat") or 0),
        "vat": float(item.get("vat") or 0),
        "withholding_tax": float(item.get("withholdingTax") or item.get("withholding_tax") or 0),
        "total": float(item.get("total") or 0),
        "document_type": item.get("documentType") or "-",
        "submitter_name": item.get("submitterName") or "-",
        "sheet_name": item.get("sheetName") or "-",
        "row": item.get("row") or "-",
    }


def render_substitute_receipt_image(data: dict[str, Any]) -> Path:
    from PIL import Image, ImageDraw

    out_dir = reply_image_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"substitute_receipt_{data.get('row', '-')}_{uuid.uuid4().hex[:8]}.png"
    width, height = 1280, 1700
    image = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    title_font = find_font(42)
    header_font = find_font(30)
    label_font = find_font(25)
    small_font = find_font(21)

    draw.rectangle((38, 38, width - 38, height - 38), outline="#111827", width=3)
    draw.text((width // 2 - 250, 70), "เนเธเธฃเธฑเธเธฃเธญเธเนเธ—เธเนเธเน€เธชเธฃเนเธเธฃเธฑเธเน€เธเธดเธ", fill="#111827", font=title_font)
    draw.text((width // 2 - 280, 125), "เธเธฃเธ“เธตเนเธกเนเธกเธตเนเธเน€เธชเธฃเนเธเธฃเธฑเธเน€เธเธดเธ/เนเธ”เนเธฃเธฑเธเน€เธเธตเธขเธเธเธดเธฅเธซเธฃเธทเธญเธเธดเธฅเน€เธเธดเธเธชเธ”", fill="#374151", font=small_font)
    draw.line((80, 175, width - 80, 175), fill="#111827", width=2)

    y = 210
    rows = [
        ("เธงเธฑเธเธ—เธตเนเธเธฑเธ”เธ—เธณเนเธเนเธ—เธ", dt.date.today().isoformat()),
        ("เธงเธฑเธเธ—เธตเนเธ•เธฒเธกเน€เธญเธเธชเธฒเธฃ/เธงเธฑเธเธ—เธตเนเธเนเธฒเธข", str(data.get("date") or "-")),
        ("เธญเนเธฒเธเธญเธดเธ Google Sheet", f"{data.get('sheet_name')} Row {data.get('row')}"),
        ("เธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเน€เธ”เธดเธก", data.get("document_type") or "-"),
        ("เน€เธฅเธเธ—เธตเนเธเธดเธฅ/เน€เธญเธเธชเธฒเธฃ", data.get("invoice_no") or "-"),
        ("เธเธทเนเธญเธฃเนเธฒเธ/เธเธนเนเธเนเธฒ/เธเธนเนเธฃเธฑเธเน€เธเธดเธ", data.get("vendor") or "-"),
        ("เธซเธกเธงเธ”เธเนเธฒเนเธเนเธเนเธฒเธข", data.get("category") or "-"),
        ("เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธเนเธฒเนเธเนเธเนเธฒเธข", data.get("description") or "-"),
        ("เธเธนเนเธเธณเธชเนเธเน€เธญเธเธชเธฒเธฃ", data.get("submitter_name") or "-"),
    ]
    for label, value in rows:
        draw.text((85, y), label, fill="#374151", font=label_font)
        wrapped = wrap_text(str(value), 48)
        line_y = y
        for line in wrapped:
            draw.text((430, line_y), line, fill="#111827", font=label_font)
            line_y += 32
        y += max(48, len(wrapped) * 34)

    y += 25
    draw.rounded_rectangle((80, y, width - 80, y + 190), radius=12, fill="#F8FAFC", outline="#CBD5E1")
    draw.text((110, y + 28), "เธขเธญเธ”เธเนเธญเธ VAT", fill="#374151", font=header_font)
    draw.text((760, y + 28), f"{float(data.get('before_vat') or 0):,.2f} เธเธฒเธ—", fill="#111827", font=header_font)
    draw.text((110, y + 82), "VAT", fill="#374151", font=header_font)
    draw.text((760, y + 82), f"{float(data.get('vat') or 0):,.2f} เธเธฒเธ—", fill="#111827", font=header_font)
    draw.text((110, y + 136), "เธขเธญเธ”เธฃเธงเธก", fill="#111827", font=header_font)
    draw.text((760, y + 136), f"{float(data.get('total') or 0):,.2f} เธเธฒเธ—", fill="#111827", font=header_font)

    y += 240
    note_lines = [
        "เธเนเธฒเธเน€เธเนเธฒเธเธญเธฃเธฑเธเธฃเธญเธเธงเนเธฒเนเธ”เนเธเนเธฒเธขเน€เธเธดเธเธ•เธฒเธกเธฃเธฒเธขเธเธฒเธฃเธเนเธฒเธเธ•เนเธเธเธฃเธดเธ เนเธฅเธฐเนเธกเนเธชเธฒเธกเธฒเธฃเธ–เน€เธฃเธตเธขเธ/เธฃเธฑเธเนเธเน€เธชเธฃเนเธเธฃเธฑเธเน€เธเธดเธ",
        "เธซเธฃเธทเธญเน€เธญเธเธชเธฒเธฃเธ เธฒเธฉเธตเธ—เธตเนเธชเธกเธเธนเธฃเธ“เนเธเธฒเธเธเธนเนเธฃเธฑเธเน€เธเธดเธเนเธ”เน เธเธถเธเธเธฑเธ”เธ—เธณเนเธเธฃเธฑเธเธฃเธญเธเนเธ—เธเนเธเน€เธชเธฃเนเธเธฃเธฑเธเน€เธเธดเธเธเธเธฑเธเธเธตเน",
        "เน€เธเธทเนเธญเนเธเนเน€เธเนเธเธซเธฅเธฑเธเธเธฒเธเธเธฃเธฐเธเธญเธเธเธฒเธฃเธเธฑเธเธ—เธถเธเธเนเธฒเนเธเนเธเนเธฒเธข เนเธเธฃเธ”เนเธเธเธซเธฅเธฑเธเธเธฒเธเธเธฒเธฃเธเนเธฒเธขเน€เธเธดเธ/เธฃเธนเธเธเธดเธฅเน€เธ”เธดเธกเธ—เธธเธเธเธฃเธฑเนเธ",
        "เธซเธกเธฒเธขเน€เธซเธ•เธธ: เธเธฃเธ“เธตเนเธเนเธขเธทเนเธเธ เธฒเธฉเธต เธเธงเธฃเนเธซเนเธเธนเนเธ—เธณเธเธฑเธเธเธตเธซเธฃเธทเธญเธ—เธตเนเธเธฃเธถเธเธฉเธฒเธ เธฒเธฉเธตเธ•เธฃเธงเธเธชเธญเธเธเธงเธฒเธกเน€เธซเธกเธฒเธฐเธชเธกเธเนเธญเธเธขเธทเนเธ",
    ]
    for line in note_lines:
        draw.text((90, y), line, fill="#111827", font=small_font)
        y += 34

    y += 70
    signature_blocks = [
        ("เธเธนเนเธเนเธฒเธขเน€เธเธดเธ/เธเธนเนเธเธญเน€เธเธดเธ", data.get("submitter_name") or ""),
        ("เธเธนเนเธ•เธฃเธงเธเธชเธญเธ/เธเธนเนเธญเธเธธเธกเธฑเธ•เธด", ""),
        ("เธเธนเนเธฃเธฑเธเน€เธเธดเธ", data.get("vendor") or ""),
    ]
    block_width = 350
    x_positions = [90, 465, 840]
    for x, (label, name) in zip(x_positions, signature_blocks):
        draw.line((x, y, x + block_width, y), fill="#111827", width=2)
        draw.text((x + 45, y + 15), label, fill="#374151", font=small_font)
        draw.text((x + 35, y + 48), f"({name or '________________'})", fill="#111827", font=small_font)
        draw.text((x + 65, y + 82), "เธงเธฑเธเธ—เธตเน ____/____/______", fill="#374151", font=small_font)

    draw.text((90, height - 90), "เน€เธญเธเธชเธฒเธฃเธชเธฃเนเธฒเธเธเธฒเธเธฃเธฐเธเธ LINE VAT Bot เนเธฅเธฐเธเนเธญเธกเธนเธฅเนเธ Google Sheet", fill="#64748B", font=small_font)
    image.save(path, quality=94)
    return path


def render_substitute_receipt_pdf(image_path: Path) -> Path:
    from PIL import Image

    pdf_path = image_path.with_suffix(".pdf")
    with Image.open(image_path) as image:
        image.convert("RGB").save(pdf_path, "PDF", resolution=150.0)
    return pdf_path


def substitute_receipt_messages(matches: list[dict[str, Any]], public_base_url: str, line_user_id: str = "") -> list[dict[str, Any]]:
    data = substitute_data_from_match(matches[0])
    image_path = render_substitute_receipt_image(data)
    pdf_path = render_substitute_receipt_pdf(image_path)
    image_url = public_file_url(public_base_url, image_path)
    pdf_url = public_file_url(public_base_url, pdf_path)
    save_note = "เธเธฑเธเธ—เธถเธเธเธฃเธฐเธงเธฑเธ•เธดเนเธเนเธ—เธเธฅเธ Google Sheet เนเธฅเนเธง"
    try:
        result = save_substitute_receipt_to_google(data, image_path, pdf_path, line_user_id)
        save_note = f"เธเธฑเธเธ—เธถเธเธเธฃเธฐเธงเธฑเธ•เธดเนเธเนเธ—เธเธฅเธ Google Sheet เนเธฅเนเธง ({result.get('sheetName')} Row {result.get('row')})"
    except Exception as exc:
        runtime_log(f"Save substitute receipt record failed: {exc}")
        save_note = f"เธชเธฃเนเธฒเธเนเธเนเธ—เธเธชเธณเน€เธฃเนเธ เนเธ•เนเธเธฑเธเธ—เธถเธเธเธฃเธฐเธงเธฑเธ•เธดเธฅเธ Google Sheet เนเธกเนเธชเธณเน€เธฃเนเธเธเธฑเนเธงเธเธฃเธฒเธง ({exc})"
    return [
        text_message(
            "เธชเธฃเนเธฒเธเนเธเนเธ—เธเน€เธฃเธตเธขเธเธฃเนเธญเธข\n"
            f"เธญเนเธฒเธเธญเธดเธ {data.get('sheet_name')} Row {data.get('row')}\n"
            f"{save_note}\n"
            "เนเธเธฃเธ”เธ•เธฃเธงเธเธชเธญเธเนเธฅเธฐเนเธซเนเธเธนเนเธกเธตเธญเธณเธเธฒเธเธฅเธเธเธฒเธกเธเนเธญเธเนเธเนเน€เธเนเธเน€เธญเธเธชเธฒเธฃเธเธฃเธฐเธเธญเธเธเธฑเธเธเธต\n"
            f"PDF เธชเธณเธซเธฃเธฑเธเธเธดเธกเธเน: {pdf_url}"
        ),
        image_message(image_url),
    ]


def public_file_url(base_url: str, path: Path) -> str:
    return f"{base_url.rstrip('/')}/files/{path.name}"


def download_binary(url: str, timeout: int = 45) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "LineExpenseBot/2.0"})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def render_pdf_first_page_to_jpeg(pdf_bytes: bytes, label: str = "schedule") -> Path:
    import fitz
    from PIL import Image

    out_dir = reply_image_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_label = re.sub(r"[^A-Za-z0-9_-]+", "_", label).strip("_") or "schedule"
    path = out_dir / f"{safe_label}_{uuid.uuid4().hex[:8]}.jpg"
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        page = document.load_page(0)
        pixmap = page.get_pixmap(matrix=fitz.Matrix(2.2, 2.2), alpha=False)
        png_bytes = pixmap.tobytes("png")
    finally:
        document.close()
    image = Image.open(io.BytesIO(png_bytes)).convert("RGB")
    image.save(path, "JPEG", quality=92, optimize=True)
    return path


def to_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        amount = normalize_amount(value)
        if amount is not None:
            return float(amount)
    return 0.0


def transaction_row_data(sheet, row: int) -> dict[str, Any]:
    data = {
        "date": sheet.cell(row=row, column=1).value,
        "transaction_type": sheet.cell(row=row, column=2).value,
        "invoice_no": sheet.cell(row=row, column=3).value,
        "document_type": sheet.cell(row=row, column=18).value,
        "vendor": sheet.cell(row=row, column=4).value,
        "description": sheet.cell(row=row, column=5).value,
        "category": sheet.cell(row=row, column=6).value,
        "before_vat": sheet.cell(row=row, column=7).value or 0,
        "confidence": sheet.cell(row=row, column=14).value or 0,
    }
    before_vat = to_float(data["before_vat"])
    vat_rate = float(CONFIG.get("vat_rate", 0.07))
    data["vat"] = round(before_vat * vat_rate, 2)
    data["total"] = round(before_vat + data["vat"], 2)
    return data


def find_bill_in_transactions(bill_no: str) -> tuple[str, int, dict[str, Any]] | None:
    workbook_path = resolve_path(CONFIG["workbook"])
    wb = load_workbook(workbook_path, data_only=False)
    if "Transactions_12M" not in wb.sheetnames:
        return None
    sheet = wb["Transactions_12M"]
    target = bill_no.strip().lower()
    for row in range(4, sheet.max_row + 1):
        current = sheet.cell(row=row, column=3).value
        if current and str(current).strip().lower() == target:
            return "Transactions_12M", row, transaction_row_data(sheet, row)
    return None


def search_transactions(query: str, max_results: int = 5) -> list[tuple[str, int, dict[str, Any]]]:
    workbook_path = resolve_path(CONFIG["workbook"])
    wb = load_workbook(workbook_path, data_only=False)
    if "Transactions_12M" not in wb.sheetnames:
        return []
    sheet = wb["Transactions_12M"]
    q = query.strip()
    q_lower = q.lower()
    q_amount = normalize_amount(q)
    results: list[tuple[int, str, int, dict[str, Any]]] = []

    for row in range(4, sheet.max_row + 1):
        if not sheet.cell(row=row, column=1).value or not sheet.cell(row=row, column=2).value:
            continue
        data = transaction_row_data(sheet, row)
        if not data.get("transaction_type") in {"Revenue", "Expense"}:
            continue
        invoice = str(data.get("invoice_no") or "").lower()
        vendor = str(data.get("vendor") or "").lower()
        before_vat = float(data.get("before_vat") or 0)
        total = float(data.get("total") or 0)
        score = 0
        if q_lower and invoice and q_lower == invoice:
            score += 100
        if q_lower and invoice and q_lower in invoice:
            score += 60
        if q_lower and vendor and q_lower in vendor:
            score += 50
        if q_amount is not None:
            if abs(total - q_amount) <= 1:
                score += 80
            elif abs(before_vat - q_amount) <= 1:
                score += 55
        if score:
            results.append((score, "Transactions_12M", row, data))

    results.sort(key=lambda item: (item[0], item[2]), reverse=True)
    return [(sheet_name, row, data) for _, sheet_name, row, data in results[:max_results]]


def find_transaction_by_row(row: int) -> tuple[str, int, dict[str, Any]] | None:
    workbook_path = resolve_path(CONFIG["workbook"])
    wb = load_workbook(workbook_path, data_only=False)
    if "Transactions_12M" not in wb.sheetnames:
        return None
    sheet = wb["Transactions_12M"]
    if row < 4 or row > sheet.max_row:
        return None
    if not sheet.cell(row=row, column=1).value or not sheet.cell(row=row, column=2).value:
        return None
    return "Transactions_12M", row, transaction_row_data(sheet, row)


def format_lookup_results(results: list[tuple[str, int, dict[str, Any]]], query: str) -> str:
    lines = [f"เธเธเธซเธฅเธฒเธขเธฃเธฒเธขเธเธฒเธฃเธเธฒเธเธเธณเธเนเธ: {query}", "เธเธฃเธธเธ“เธฒเธเธดเธกเธเน Row เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃ เน€เธเนเธ Row 170"]
    for sheet_name, row, data in results:
        lines.append(
            f"Row {row}: {data.get('date')} | {data.get('vendor') or '-'} | "
            f"{data.get('invoice_no') or '-'} | Total {float(data.get('total') or 0):,.2f}"
        )
    return "\n".join(lines)


def reset_transaction_formulas(sheet, row: int) -> None:
    sheet.cell(row=row, column=8, value=f'=IF(A{row}="","",Settings!$B$4)')
    sheet.cell(row=row, column=9, value=f'=IF(G{row}="","",IF(OR(B{row}="Revenue",K{row}="Yes"),ROUND(G{row}*H{row},2),0))')
    sheet.cell(row=row, column=10, value=f'=IF(G{row}="","",G{row}+ROUND(G{row}*H{row},2))')
    sheet.cell(row=row, column=12, value=f'=IF(A{row}="","",DATE(YEAR(A{row}),MONTH(A{row}),1))')
    sheet.cell(row=row, column=15, value=f'=IF(B{row}="Revenue",G{row},0)')
    sheet.cell(row=row, column=16, value=f'=IF(B{row}="Expense",G{row},0)')


def clear_transaction_row(sheet, row: int) -> None:
    # Clear user/OCR-entered cells but keep the formula columns ready for reuse.
    for col in (1, 2, 3, 4, 5, 6, 7, 11, 13, 14, 17, 18):
        sheet.cell(row=row, column=col).value = None
    reset_transaction_formulas(sheet, row)


def clear_legacy_expense_row(sheet, row: int) -> None:
    for col in (1, 2, 3, 4, 5, 6, 10, 12, 13, 14, 15):
        sheet.cell(row=row, column=col).value = None
    sheet.cell(row=row, column=7, value=f'=IF(A{row}="","",Settings!$B$4)')
    sheet.cell(row=row, column=8, value=f'=IF(F{row}="","",IF(J{row}="Yes",ROUND(F{row}*G{row},2),0))')
    sheet.cell(row=row, column=9, value=f'=IF(F{row}="","",F{row}+ROUND(F{row}*G{row},2))')
    sheet.cell(row=row, column=11, value=f'=IF(A{row}="","",DATE(YEAR(A{row}),MONTH(A{row}),1))')


def delete_bill_from_excel(bill_no: str, line_user_id: str = "") -> str:
    bill_no = bill_no.strip()
    if not bill_no:
        return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเน€เธฅเธเธ—เธตเนเธเธดเธฅ เน€เธเนเธ เนเธเนเนเธเธเธดเธฅ+INV001"

    workbook_path = resolve_path(CONFIG["workbook"])
    wb = load_workbook(workbook_path)
    deleted: list[tuple[str, int]] = []

    if "Transactions_12M" in wb.sheetnames:
        sheet = wb["Transactions_12M"]
        for row in range(5, sheet.max_row + 1):
            value = str(sheet.cell(row=row, column=3).value or "").strip()
            if value and value.lower() == bill_no.lower():
                clear_transaction_row(sheet, row)
                deleted.append(("Transactions_12M", row))

    if "Expenses" in wb.sheetnames:
        sheet = wb["Expenses"]
        for row in range(4, sheet.max_row + 1):
            value = str(sheet.cell(row=row, column=2).value or "").strip()
            if value and value.lower() == bill_no.lower():
                clear_legacy_expense_row(sheet, row)
                deleted.append(("Expenses", row))

    log = wb["Import_Log"]
    log_row = next_empty_row(log)
    log.cell(row=log_row, column=1, value=dt.datetime.now())
    log.cell(row=log_row, column=2, value="")
    log.cell(row=log_row, column=3, value="Deleted" if deleted else "DeleteNotFound")
    log.cell(row=log_row, column=4, value=dt.date.today())
    log.cell(row=log_row, column=5, value=bill_no)
    log.cell(row=log_row, column=6, value=0)
    log.cell(row=log_row, column=7, value=0)
    log.cell(row=log_row, column=8, value=f"LINE delete command bill_no={bill_no} rows={deleted} LINE user: {line_user_id}")

    wb.save(workbook_path)

    if not deleted:
        runtime_log(f"Delete command: bill_no={bill_no!r} not found")
        return f"เนเธกเนเธเธเน€เธฅเธเธ—เธตเนเธเธดเธฅ: {bill_no}\nเธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเน€เธฅเธเธ—เธตเนเน€เธญเธเธชเธฒเธฃเนเธเธเธตเธ• Transactions_12M"

    runtime_log(f"Delete command: bill_no={bill_no!r} cleared rows={deleted}")
    row_text = ", ".join(f"{sheet}!row {row}" for sheet, row in deleted)
    return (
        "==== เธเธดเธฅเธขเธเน€เธฅเธดเธ ====\n"
        "เธชเธ–เธฒเธเธฐ: เธขเธเน€เธฅเธดเธเธเธดเธฅเน€เธฃเธตเธขเธเธฃเนเธญเธข\n"
        f"เน€เธฅเธเธ—เธตเนเธเธดเธฅ: {bill_no}\n"
        f"เธ•เธณเนเธซเธเนเธ: {row_text}\n"
        "เธชเธฒเธกเธฒเธฃเธ–เธชเนเธเธฃเธนเธเธเธดเธฅเนเธซเธกเนเธซเธฃเธทเธญเธเธฃเธญเธเนเธซเธกเนเนเธ”เนเน€เธฅเธข"
    )


def delete_excel_row(row_text: str, line_user_id: str = "") -> str:
    row_text = row_text.strip()
    if not row_text.isdigit():
        return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเน€เธฅเธเนเธ–เธง เน€เธเนเธ Delete Row 173"

    row = int(row_text)
    workbook_path = resolve_path(CONFIG["workbook"])
    wb = load_workbook(workbook_path)

    if "Transactions_12M" not in wb.sheetnames:
        return "เนเธกเนเธเธเธเธตเธ• Transactions_12M เนเธเนเธเธฅเน Excel"

    sheet = wb["Transactions_12M"]
    if row < 5 or row > sheet.max_row:
        return f"เน€เธฅเธเนเธ–เธง {row} เธญเธขเธนเนเธเธญเธเธเนเธงเธเธเนเธญเธกเธนเธฅเธ—เธตเนเธฅเธเนเธ”เน"

    current_date = sheet.cell(row=row, column=1).value
    current_type = sheet.cell(row=row, column=2).value
    current_doc = sheet.cell(row=row, column=3).value
    current_party = sheet.cell(row=row, column=4).value
    current_amount = sheet.cell(row=row, column=7).value

    if not any([current_date, current_type, current_doc, current_party, current_amount]):
        return f"Row {row} เนเธกเนเธกเธตเธเนเธญเธกเธนเธฅเนเธซเนเธฅเธ"

    clear_transaction_row(sheet, row)

    log = wb["Import_Log"]
    log_row = next_empty_row(log)
    log.cell(row=log_row, column=1, value=dt.datetime.now())
    log.cell(row=log_row, column=2, value="")
    log.cell(row=log_row, column=3, value="DeletedRow")
    log.cell(row=log_row, column=4, value=dt.date.today())
    log.cell(row=log_row, column=5, value=f"Row {row}")
    log.cell(row=log_row, column=6, value=current_amount or 0)
    log.cell(row=log_row, column=7, value=0)
    log.cell(
        row=log_row,
        column=8,
        value=(
            f"LINE Delete Row command row={row} "
            f"old_doc={current_doc} old_party={current_party} LINE user: {line_user_id}"
        ),
    )

    wb.save(workbook_path)
    runtime_log(f"Delete Row command: cleared Transactions_12M!row {row}")
    return (
        "==== เธเธดเธฅเธขเธเน€เธฅเธดเธ ====\n"
        "เธชเธ–เธฒเธเธฐ: เธขเธเน€เธฅเธดเธเธเธดเธฅเน€เธฃเธตเธขเธเธฃเนเธญเธข\n"
        f"Row: {row}\n"
        f"เน€เธฅเธเธ—เธตเนเน€เธญเธเธชเธฒเธฃเน€เธ”เธดเธก: {current_doc or '-'}\n"
        f"เธเธทเนเธญเน€เธ”เธดเธก: {current_party or '-'}\n"
        "เธชเธฒเธกเธฒเธฃเธ–เธชเนเธเธฃเธนเธเธเธดเธฅเนเธซเธกเนเธซเธฃเธทเธญเธเธฃเธญเธเนเธซเธกเนเนเธ”เนเน€เธฅเธข"
    )


def append_transaction_to_excel(image_path: Path, data: dict[str, Any], status: str, message: str, line_user_id: str = "") -> tuple[str, int]:
    workbook_path = resolve_path(CONFIG["workbook"])
    wb = load_workbook(workbook_path)
    data = apply_transaction_type_defaults(data, data.get("transaction_type", "Expense"))

    if "Transactions_12M" in wb.sheetnames:
        target_sheet = "Transactions_12M"
        sheet = wb[target_sheet]
        row = find_transaction_row(sheet, data["date"])
        append_to_transactions(sheet, row, image_path, data)
        if data.get("transaction_type") == "Expense" and "Expenses" in wb.sheetnames:
            legacy_row = next_empty_row(wb["Expenses"])
            append_to_legacy_expenses(wb["Expenses"], legacy_row, image_path, data)
        elif data.get("transaction_type") == "Revenue" and "Revenue" in wb.sheetnames:
            legacy_row = next_empty_row(wb["Revenue"])
            append_to_legacy_revenue(wb["Revenue"], legacy_row, data)
    else:
        target_sheet = "Expenses"
        sheet = wb[target_sheet]
        row = next_empty_row(sheet)
        append_to_legacy_expenses(sheet, row, image_path, data)

    append_import_log(wb["Import_Log"], image_path, data, status, message, line_user_id)
    wb.save(workbook_path)
    return target_sheet, row


def append_expense_to_excel(image_path: Path, data: dict[str, Any], status: str, message: str, line_user_id: str = "") -> tuple[str, int]:
    data = apply_transaction_type_defaults(data, "Expense")
    return append_transaction_to_excel(image_path, data, status, message, line_user_id)


def build_google_sheet_payload(data: dict[str, Any], image_path: Path, line_user_id: str, public_base_url: str, message: str) -> dict[str, Any]:
    before_vat = float(data.get("before_vat") or 0)
    vat_rate = float(CONFIG.get("vat_rate", 0.07))
    vat = float(data.get("vat") or round(before_vat * vat_rate, 2))
    total = float(data.get("total") or round(before_vat + vat, 2))
    image_url = ""
    image_file_name = ""
    image_data = ""
    image_mime_type = ""
    if image_path and image_path.name and public_base_url:
        image_file_name = image_path.name
        image_mime_type = mimetypes.guess_type(str(image_path))[0] or "image/jpeg"
        image_data = base64.b64encode(image_path.read_bytes()).decode("ascii")
    date_value = data.get("date") or dt.date.today()
    if isinstance(date_value, (dt.datetime, dt.date)):
        date_text = date_value.isoformat()[:10]
        month_text = f"{date_value.year:04d}-{date_value.month:02d}"
    else:
        date_text = str(date_value)
        month_text = str(date_value)[:7]
    document_paths = [str(path) for path in data.get("document_paths") or []]
    return {
        "secret": CONFIG.get("google_apps_script_secret", ""),
        "date": date_text,
        "type": data.get("transaction_type", "Expense"),
        "transaction_type": data.get("transaction_type", "Expense"),
        "invoiceNo": normalize_invoice_no(data.get("invoice_no")),
        "invoice_no": normalize_invoice_no(data.get("invoice_no")),
        "vendor": data.get("vendor", ""),
        "description": data.get("description", ""),
        "category": data.get("category", ""),
        "beforeVat": before_vat,
        "before_vat": before_vat,
        "vatRate": vat_rate,
        "vat_rate": vat_rate,
        "vat": vat,
        "withholdingTax": float(data.get("withholding_tax") or data.get("withholdingTax") or 0),
        "withholding_tax": float(data.get("withholding_tax") or data.get("withholdingTax") or 0),
        "total": total,
        "claimable": data.get("claimable", "Yes"),
        "month": month_text,
        "imageUrl": image_file_name or image_url,
        "image_url": image_file_name or image_url,
        "imageFileName": image_file_name,
        "image_file_name": image_file_name,
        "imageMimeType": image_mime_type,
        "image_mime_type": image_mime_type,
        "imageData": image_data,
        "image_data": image_data,
        "documentImages": document_paths,
        "document_images": document_paths,
        "documentCount": len(document_paths) or (1 if image_file_name else 0),
        "document_count": len(document_paths) or (1 if image_file_name else 0),
        "confidence": data.get("confidence", ""),
        "rawText": data.get("raw_text", ""),
        "raw_text": data.get("raw_text", ""),
        "documentType": data.get("document_type", ""),
        "document_type": data.get("document_type", ""),
        "submitterName": data.get("submitter_name", ""),
        "submitter_name": data.get("submitter_name", ""),
        "lineUserId": line_user_id,
        "line_user_id": line_user_id,
        "message": message,
    }


def send_to_google_sheet(data: dict[str, Any], image_path: Path, line_user_id: str, public_base_url: str) -> dict[str, Any] | None:
    url = CONFIG.get("google_apps_script_url")
    if not url:
        return None
    payload = build_google_sheet_payload(data, image_path, line_user_id, public_base_url, "Confirmed LINE OCR import")
    result = post_json(url, payload)
    runtime_log(f"Google Apps Script response: {result}")
    if result.get("status") != "ok":
        raise RuntimeError(f"Google Apps Script save failed: {result}")
    return result


def google_sheet_action(action: str, **kwargs: Any) -> dict[str, Any]:
    url = CONFIG.get("google_apps_script_url")
    if not url:
        raise RuntimeError("missing google_apps_script_url")
    payload = {"secret": CONFIG.get("google_apps_script_secret", ""), "action": action}
    payload.update(kwargs)
    result = post_json(url, payload)
    runtime_log(f"Google Apps Script action {action} response: {result}")
    if result.get("status") != "ok":
        raise RuntimeError(f"Google Apps Script action {action} failed: {result}")
    return result


def search_google_sheet_by_total(total: Any) -> list[dict[str, Any]]:
    result = google_sheet_action("searchByTotal", total=to_float(total))
    return list(result.get("matches") or [])


def search_stock_product(branch: str, query: str) -> list[dict[str, Any]]:
    source = str(CONFIG.get("stock_source") or "auto").lower()
    if source in {"qashier", "auto"}:
        try:
            return search_qashier_stock(branch, query)
        except Exception as exc:
            runtime_log(f"Qashier stock search skipped/failed: {exc}")
            if source == "qashier":
                raise

    result = google_sheet_action("searchStock", branch=branch, query=query)
    matches = list(result.get("matches") or [])
    for item in matches:
        item.setdefault("source", "Google Sheet")
    return matches


def qashier_config() -> dict[str, Any]:
    return dict(CONFIG.get("qashier") or {})


def qashier_env(config_key: str, default: str = "") -> str:
    env_name = qashier_config().get(config_key)
    if not env_name:
        return default
    return os.getenv(str(env_name), default)


def search_qashier_stock(branch: str, query: str) -> list[dict[str, Any]]:
    cfg = qashier_config()
    url = qashier_env("inventory_url_env")
    token = qashier_env("api_token_env")
    store_id = qashier_env("store_id_env")
    if not url or not token:
        raise RuntimeError("missing Qashier inventory API URL or token")

    params = {
        "q": query,
        "query": query,
        "keyword": query,
        "barcode": query,
        "branch": branch,
        "storeId": store_id,
        "store_id": store_id,
    }
    clean_params = {k: v for k, v in params.items() if v}
    separator = "&" if "?" in url else "?"
    endpoint = f"{url}{separator}{urllib.parse.urlencode(clean_params)}"
    request = urllib.request.Request(
        endpoint,
        headers={
            "Accept": "application/json",
            "Authorization": f"Bearer {token}",
            "User-Agent": "LineVatBot/2.0",
        },
        method="GET",
    )
    timeout = int(cfg.get("timeout_seconds") or 20)
    with urllib.request.urlopen(request, timeout=timeout) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return normalize_qashier_stock_items(payload, branch)


def normalize_qashier_stock_items(payload: Any, branch: str) -> list[dict[str, Any]]:
    if isinstance(payload, dict):
        raw_items = (
            payload.get("items")
            or payload.get("products")
            or payload.get("inventory")
            or payload.get("results")
            or payload.get("data")
            or []
        )
    else:
        raw_items = payload
    if isinstance(raw_items, dict):
        raw_items = (
            raw_items.get("items")
            or raw_items.get("products")
            or raw_items.get("inventory")
            or raw_items.get("results")
            or []
        )
    if not isinstance(raw_items, list):
        return []

    def pick(item: dict[str, Any], *keys: str) -> Any:
        for key in keys:
            value = item.get(key)
            if value not in (None, ""):
                return value
        return ""

    normalized: list[dict[str, Any]] = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        stock_value = pick(
            item,
            "stock",
            "quantity",
            "qty",
            "availableQty",
            "available_quantity",
            "inventoryQuantity",
        )
        normalized.append(
            {
                "name": pick(item, "name", "productName", "product_name", "title"),
                "price": pick(item, "price", "sellingPrice", "selling_price", "retailPrice"),
                "barcode": pick(item, "barcode", "sku", "code", "productCode"),
                "stock": stock_value,
                "branch": pick(item, "branch", "storeName", "store_name") or branch,
                "source": "Qashier HQ",
            }
        )
    return normalized


def format_stock_results(branch: str, query: str, matches: list[dict[str, Any]]) -> str:
    if not matches:
        return (
            f"เนเธกเนเธเธเธชเธดเธเธเนเธฒเนเธเธชเธฒเธเธฒ {branch}\n"
            f"เธเธณเธเนเธ: {query}\n\n"
            "เธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเธชเธญเธเธเธทเนเธญเธชเธดเธเธเนเธฒ/เธเธฒเธฃเนเนเธเนเธ”/SKU เนเธฅเนเธงเธเนเธเธซเธฒเธญเธตเธเธเธฃเธฑเนเธเธเนเธฐ"
        )
    lines = [
        f"เธเธเธชเธดเธเธเนเธฒเนเธเธชเธฒเธเธฒ {branch}",
        f"เธเธณเธเนเธ: {query}",
        "",
    ]
    for idx, item in enumerate(matches[:10], 1):
        lines.extend(
            [
                f"{idx}. เธเธทเนเธญเธชเธดเธเธเนเธฒ: {item.get('name') or '-'}",
                f"เธฃเธฒเธเธฒเธชเธดเธเธเนเธฒ: {item.get('price') or '-'}",
                f"เธเธฒเธฃเนเนเธเนเธ”: {item.get('barcode') or '-'}",
                "",
            ]
        )
    if len(matches) > 10:
        lines.append(f"เนเธชเธ”เธ 10 เธฃเธฒเธขเธเธฒเธฃเนเธฃเธเธเธฒเธเธ—เธฑเนเธเธซเธกเธ” {len(matches)} เธฃเธฒเธขเธเธฒเธฃ")
    return "\n".join(lines).strip()


def parse_stock_queries(text: str, limit: int = 10) -> tuple[list[str], int]:
    raw_parts = re.split(r"[\n,;]+", text)
    queries: list[str] = []
    for part in raw_parts:
        query = part.strip()
        if query and query not in queries:
            queries.append(query)
    return queries[:limit], len(queries)


def format_stock_check_results(branch: str, queries: list[str], results: dict[str, list[dict[str, Any]]], total_entered: int) -> str:
    lines = [
        f"เธเธฅเธ•เธฃเธงเธเธชเธญเธเธชเธ•เนเธญเธ เธชเธฒเธเธฒ {branch}",
        f"เธ•เธฃเธงเธเธชเธญเธ {len(queries)} เธฃเธฒเธขเธเธฒเธฃ" + (" เธเธฒเธ 10 เธฃเธฒเธขเธเธฒเธฃเนเธฃเธ" if total_entered > 10 else ""),
        "",
    ]
    for idx, query in enumerate(queries, 1):
        matches = results.get(query) or []
        lines.append(f"{idx}. เธเธณเธเนเธ/เธเธฒเธฃเนเนเธเนเธ”: {query}")
        if not matches:
            lines.extend(["เนเธกเนเธเธเธชเธดเธเธเนเธฒ", ""])
            continue
        for item in matches[:3]:
            qty = item.get("quantity")
            lines.extend(
                [
                    f"เธเธทเนเธญเธชเธดเธเธเนเธฒ: {item.get('name') or '-'}",
                    f"เธเธณเธเธงเธเธเธเน€เธซเธฅเธทเธญ: {qty if qty not in [None, ''] else '-'}",
                    f"เธฃเธฒเธเธฒ: {item.get('price') or '-'}",
                    f"เธเธฒเธฃเนเนเธเนเธ”: {item.get('barcode') or '-'}",
                    "",
                ]
            )
        if len(matches) > 3:
            lines.extend([f"เธเธเธ—เธฑเนเธเธซเธกเธ” {len(matches)} เธฃเธฒเธขเธเธฒเธฃ เนเธชเธ”เธ 3 เธฃเธฒเธขเธเธฒเธฃเนเธฃเธ", ""])
    lines.append("เธเนเธญเธกเธนเธฅเธญเนเธฒเธเธญเธดเธเธเธฒเธ Product List/Qashier HQ เธ—เธตเนเธเธดเธเธเนเนเธงเนเนเธ Google Sheet")
    return "\n".join(lines).strip()


def delete_google_sheet_row(row: int, sheet_name: str = "") -> dict[str, Any]:
    payload: dict[str, Any] = {"row": int(row)}
    if sheet_name:
        payload["sheetName"] = sheet_name
    return google_sheet_action("deleteRow", **payload)


def update_google_sheet_document_type(row: int, document_type: str, sheet_name: str = "") -> dict[str, Any]:
    payload: dict[str, Any] = {"row": int(row), "documentType": document_type}
    if sheet_name:
        payload["sheetName"] = sheet_name
    return google_sheet_action("updateDocumentType", **payload)


def save_hr_request_to_google(data: dict[str, Any], line_user_id: str) -> dict[str, Any]:
    payload = dict(data)
    payload["lineUserId"] = line_user_id
    return google_sheet_action("saveHrRequest", data=payload)


def get_hr_schedule_link(month_offset: int = 0) -> dict[str, Any]:
    return google_sheet_action("getHrSchedule", monthOffset=month_offset)


def save_medical_certificate_to_google(request_id: str, image_path: Path, line_user_id: str) -> dict[str, Any]:
    raw = image_path.read_bytes()
    mime_type = mimetypes.guess_type(str(image_path))[0] or "image/jpeg"
    return google_sheet_action(
        "saveMedicalCertificate",
        requestId=request_id,
        lineUserId=line_user_id,
        fileName=image_path.name,
        mimeType=mime_type,
        data=base64.b64encode(raw).decode("ascii"),
    )


def save_substitute_receipt_to_google(data: dict[str, Any], image_path: Path, pdf_path: Path, line_user_id: str) -> dict[str, Any]:
    image_mime_type = mimetypes.guess_type(str(image_path))[0] or "image/png"
    payload = {
        "sheetName": data.get("sheet_name") or "",
        "transactionRow": data.get("row") or "",
        "date": data.get("date") or "",
        "documentType": data.get("document_type") or "",
        "invoiceNo": data.get("invoice_no") or "",
        "vendor": data.get("vendor") or "",
        "submitterName": data.get("submitter_name") or "",
        "category": data.get("category") or "",
        "beforeVat": data.get("before_vat") or 0,
        "vat": data.get("vat") or 0,
        "withholdingTax": data.get("withholding_tax") or data.get("withholdingTax") or 0,
        "total": data.get("total") or 0,
        "imageUrl": image_path.name,
        "imageFileName": image_path.name,
        "imageMimeType": image_mime_type,
        "imageData": base64.b64encode(image_path.read_bytes()).decode("ascii"),
        "pdfUrl": pdf_path.name,
        "lineUserId": line_user_id,
    }
    return google_sheet_action("saveSubstituteReceipt", data=payload)


def update_hr_approval_to_google(request_id: str, approved: bool, approver_line_id: str) -> dict[str, Any]:
    return google_sheet_action(
        "updateHrApproval",
        requestId=request_id,
        approved=approved,
        note=f"Updated from LINE by {approver_line_id}",
    )


def format_google_sheet_matches(matches: list[dict[str, Any]], heading: str) -> str:
    lines = [heading]
    for item in matches[:10]:
        lines.append(
            f"{item.get('sheetName') or 'Sheet'} Row {item.get('row')}: {item.get('date') or '-'} | "
            f"{item.get('type') or '-'} | {item.get('vendor') or '-'} | "
            f"เน€เธฅเธเธ—เธตเนเธเธดเธฅ {item.get('invoiceNo') or '-'} | "
            f"เธเนเธญเธ VAT {float(item.get('beforeVat') or 0):,.2f} | "
            f"เธขเธญเธ”เธฃเธงเธก {float(item.get('total') or 0):,.2f} | "
            f"เน€เธญเธเธชเธฒเธฃ {item.get('documentType') or '-'}"
        )
    return "\n".join(lines)


def process_line_event(event: dict[str, Any]) -> str | None:
    runtime_log(f"Received LINE event type={event.get('type')} message_type={event.get('message', {}).get('type')}")
    if event.get("type") != "message":
        return None
    message = event.get("message", {})
    line_user_id = event.get("source", {}).get("userId", "")

    if message.get("type") == "text":
        text = str(message.get("text") or "").strip()
        if text.startswith("เนเธเนเนเธเธเธดเธฅ+"):
            bill_no = text.split("+", 1)[1].strip()
            return delete_bill_from_excel(bill_no, line_user_id)
        delete_row_match = re.match(r"^delete\s+row\s+(\d+)$", text, flags=re.IGNORECASE)
        if delete_row_match:
            return delete_excel_row(delete_row_match.group(1), line_user_id)
        return (
            "Please send a receipt image.\n"
            "เธซเธฒเธเธ•เนเธญเธเธเธฒเธฃเธฅเธ/เนเธเนเนเธเธเธดเธฅ เนเธซเนเธเธดเธกเธเน: เนเธเนเนเธเธเธดเธฅ+เน€เธฅเธเธ—เธตเนเธเธดเธฅ\n"
            "เธซเธฒเธเธ•เนเธญเธเธเธฒเธฃเธขเธเน€เธฅเธดเธเธ•เธฒเธกเนเธ–เธง เนเธซเนเธเธดเธกเธเน: Delete Row 173"
        )

    if message.get("type") not in {"image", "file"}:
        return "Please send a receipt image."

    token = os.getenv(CONFIG["line"]["channel_access_token_env"])
    if not token:
        raise RuntimeError(f"Missing {CONFIG['line']['channel_access_token_env']}")

    try:
        runtime_log("Downloading LINE image content")
        image_path = download_line_content(message["id"], token, resolve_path(CONFIG["image_archive_dir"]))
        append_document_path_to_state(state, image_path)
        runtime_log(f"Downloaded LINE image to {image_path}")
        runtime_log("OCR started")
        text = ocr_image(image_path)
        runtime_log(f"OCR completed characters={len(text)}")
    except Exception as exc:
        runtime_log(f"OCR failed: {exc}")
        draft = blank_manual_entry("Expense")
        draft["submitter_name"] = get_line_display_name(line_user_id)
        set_user_state(
            line_user_id,
            {
                "mode": "awaiting_correction",
                "transaction_type": draft["transaction_type"],
                "image_path": str(image_path) if "image_path" in locals() else "",
                "pending_data": serialize_data(draft),
            },
        )
        return manual_entry_form(draft)
    parsed = parse_receipt_text(text, float(CONFIG.get("vat_rate", 0.07)))
    sheet_name, row = append_expense_to_excel(image_path, parsed, "Imported", "Imported from LINE OCR", line_user_id)
    runtime_log(
        f"Imported LINE receipt -> {sheet_name}!row {row} "
        f"date={parsed['date']} total={parsed['total']} vendor={parsed['vendor']!r}"
    )

    return (
        "==== เธเธดเธฅเธเธณเน€เธเนเธฒ ====\n"
        "เธชเธ–เธฒเธเธฐ: เธเธฑเธเธ—เธถเธเน€เธญเธเธชเธฒเธฃเน€เธฃเธตเธขเธเธฃเนเธญเธข\n"
        f"Sheet: {sheet_name}\n"
        f"Row: {row}\n"
        f"Date: {parsed['date']}\n"
        f"Vendor: {parsed['vendor'] or '-'}\n"
        f"Category: {parsed['category']}\n"
        f"Before VAT: {parsed['before_vat']:,.2f}\n"
        f"VAT: {parsed['vat']:,.2f}\n"
        f"Total: {parsed['total']:,.2f}\n"
        f"Confidence: {parsed['confidence']:.0%}\n"
        "Please review before filing VAT."
    )


def process_line_event_menu(event: dict[str, Any], public_base_url: str) -> str | dict[str, Any] | list[dict[str, Any]] | None:
    runtime_log(f"Received LINE event type={event.get('type')} message_type={event.get('message', {}).get('type')}")
    line_user_id = event.get("source", {}).get("userId", "")
    state = get_user_state(line_user_id)
    if event.get("type") in {"follow", "join", "memberJoined"}:
        if not state.get("mode"):
            return menu_message()
        return None
    if event.get("type") != "message":
        return None
    message = event.get("message", {})

    if message.get("type") == "text":
        text = str(message.get("text") or "").strip()
        branch_match = re.match(r"^REV_BRANCH:(.+)$", text)
        if branch_match:
            branch = branch_match.group(1).strip()
            state["mode"] = "awaiting_revenue_payment_channel"
            state["transaction_type"] = "Revenue"
            state["category"] = state.get("category") or "ยอดขายหน้าสาขา"
            state["revenue_branch"] = branch
            state.setdefault("revenue_payments", {})
            set_user_state(line_user_id, state)
            return revenue_payment_menu_message()
        payment_match = re.match(r"^REV_PAYMENT:(.+)$", text)
        if payment_match:
            channel = payment_match.group(1).strip()
            state["mode"] = "awaiting_revenue_payment_amount"
            state["pending_payment_channel"] = channel
            set_user_state(line_user_id, state)
            if channel == "อื่นๆ":
                return "กรุณาพิมพ์ชื่อช่องทางรายรับอื่นๆ พร้อมยอด เช่น Shopee: 1250 หรือ เงินสดย่อย 500 ค่ะ"
            return f"กรุณาพิมพ์ยอดของ {channel} ค่ะ เช่น 1250.00"
        doc_continue_match = re.match(r"^DOC_CONTINUE:(.+)$", text)
        if doc_continue_match:
            choice = doc_continue_match.group(1).strip()
            if choice == "ส่งเอกสารต่อ":
                state["mode"] = "awaiting_image"
                set_user_state(line_user_id, state)
                return "ส่งเอกสารเพิ่มเติมในรายการเดิมได้เลยค่ะ"
            if state.get("pending_data"):
                state["mode"] = "awaiting_confirmation"
                set_user_state(line_user_id, state)
                return confirmation_prompt(deserialize_data(state["pending_data"]))
            return menu_message()
        if state.get("mode") == "awaiting_more_documents":
            if text in {"1", "ส่งเอกสารต่อ"}:
                state["mode"] = "awaiting_image"
                set_user_state(line_user_id, state)
                return "ส่งเอกสารเพิ่มเติมในรายการเดิมได้เลยค่ะ"
            if text in {"2", "ตรวจสอบและยืนยัน"} and state.get("pending_data"):
                state["mode"] = "awaiting_confirmation"
                set_user_state(line_user_id, state)
                return confirmation_prompt(deserialize_data(state["pending_data"]))
            return continue_document_menu_message()
        if state.get("mode") == "awaiting_revenue_branch":
            valid_branches = [item[0] for item in REVENUE_BRANCHES]
            if text not in valid_branches:
                return revenue_branch_menu_message()
            state["mode"] = "awaiting_revenue_payment_channel"
            state["transaction_type"] = "Revenue"
            state["category"] = state.get("category") or "ยอดขายหน้าสาขา"
            state["revenue_branch"] = text
            state.setdefault("revenue_payments", {})
            set_user_state(line_user_id, state)
            return revenue_payment_menu_message()
        if state.get("mode") == "awaiting_revenue_payment_channel":
            valid_channels = [item[0] for item in REVENUE_PAYMENT_CHANNELS]
            if text not in valid_channels:
                return revenue_payment_menu_message()
            state["mode"] = "awaiting_revenue_payment_amount"
            state["pending_payment_channel"] = text
            set_user_state(line_user_id, state)
            if text == "อื่นๆ":
                return "กรุณาพิมพ์ชื่อช่องทางรายรับอื่นๆ พร้อมยอด เช่น Shopee: 1250 หรือ เงินสดย่อย 500 ค่ะ"
            return f"กรุณาพิมพ์ยอดของ {text} ค่ะ เช่น 1250.00"
        if state.get("mode") == "awaiting_revenue_payment_amount":
            channel = str(state.get("pending_payment_channel") or "อื่นๆ")
            amount_text = text
            if channel == "อื่นๆ":
                other_match = re.match(r"^(.+?)[\s:=]+([0-9,]+(?:\.\d+)?)$", text)
                if other_match:
                    channel = other_match.group(1).strip()
                    amount_text = other_match.group(2).strip()
            amount = normalize_amount(amount_text)
            if amount is None:
                return "กรุณาพิมพ์ยอดเป็นตัวเลขค่ะ เช่น 1250.00"
            payments = dict(state.get("revenue_payments") or {})
            payments[channel] = round(amount, 2)
            state["revenue_payments"] = payments
            state.pop("pending_payment_channel", None)
            state["mode"] = "awaiting_revenue_payment_next"
            set_user_state(line_user_id, state)
            return [
                text_message("บันทึกยอดแล้วค่ะ\n" + format_revenue_payment_summary(payments)),
                simple_choice_menu(
                    "ต้องการเพิ่มช่องทางรายรับอีกไหม",
                    "REV_PAY_NEXT",
                    [
                        ("เพิ่มช่องทาง", "#DCFCE7", "#22C55E", "1"),
                        ("ส่งเอกสาร", "#EDE9FE", "#8B5CF6", "2"),
                    ],
                    "#10B981",
                ),
            ]
        pay_next_match = re.match(r"^REV_PAY_NEXT:(.+)$", text)
        if pay_next_match:
            choice = pay_next_match.group(1).strip()
            if choice == "เพิ่มช่องทาง":
                state["mode"] = "awaiting_revenue_payment_channel"
                set_user_state(line_user_id, state)
                return revenue_payment_menu_message()
            state["mode"] = "awaiting_image"
            set_user_state(line_user_id, state)
            return ask_for_receipt_after_category(f"{state.get('category', 'ยอดขายหน้าสาขา')} - {state.get('revenue_branch', '-')}")
        if state.get("mode") == "awaiting_revenue_payment_next":
            if text in {"1", "เพิ่มช่องทาง"}:
                state["mode"] = "awaiting_revenue_payment_channel"
                set_user_state(line_user_id, state)
                return revenue_payment_menu_message()
            if text in {"2", "ส่งเอกสาร"}:
                state["mode"] = "awaiting_image"
                set_user_state(line_user_id, state)
                return ask_for_receipt_after_category(f"{state.get('category', 'ยอดขายหน้าสาขา')} - {state.get('revenue_branch', '-')}")
            return simple_choice_menu(
                "ต้องการเพิ่มช่องทางรายรับอีกไหม",
                "REV_PAY_NEXT",
                [
                    ("เพิ่มช่องทาง", "#DCFCE7", "#22C55E", "1"),
                    ("ส่งเอกสาร", "#EDE9FE", "#8B5CF6", "2"),
                ],
                "#10B981",
            )
        category_match = re.match(r"^ACCT_CATEGORY:(Revenue|Expense):(.+)$", text, flags=re.IGNORECASE)
        if category_match:
            transaction_type = "Revenue" if category_match.group(1).lower() == "revenue" else "Expense"
            category = category_match.group(2).strip()
            if transaction_type == "Revenue" and category == "ยอดขายหน้าสาขา":
                set_user_state(
                    line_user_id,
                    {
                        "mode": "awaiting_revenue_branch",
                        "transaction_type": "Revenue",
                        "category": category,
                    },
                )
                return revenue_branch_menu_message()
            if category == "อื่นๆ":
                set_user_state(
                    line_user_id,
                    {
                        "mode": "awaiting_account_other_category",
                        "transaction_type": transaction_type,
                    },
                )
                return "กรุณาพิมพ์คำอธิบายหมวดหมู่อื่นๆ ที่ต้องการบันทึกค่ะ"
            set_user_state(
                line_user_id,
                {
                    "mode": "awaiting_image",
                    "transaction_type": transaction_type,
                    "category": category,
                },
            )
            return ask_for_receipt_after_category(category)
        if state.get("mode") == "awaiting_account_category":
            transaction_type = state.get("transaction_type") or "Expense"
            valid_categories = [item[0] for item in (REVENUE_CATEGORIES if transaction_type == "Revenue" else EXPENSE_CATEGORIES)]
            if text not in valid_categories:
                return category_menu_message(transaction_type)
            if text == "อื่นๆ":
                state["mode"] = "awaiting_account_other_category"
                set_user_state(line_user_id, state)
                return "กรุณาพิมพ์คำอธิบายหมวดหมู่อื่นๆ ที่ต้องการบันทึกค่ะ"
            if transaction_type == "Revenue" and text == "ยอดขายหน้าสาขา":
                set_user_state(
                    line_user_id,
                    {
                        "mode": "awaiting_revenue_branch",
                        "transaction_type": "Revenue",
                        "category": text,
                    },
                )
                return revenue_branch_menu_message()
            set_user_state(
                line_user_id,
                {
                    "mode": "awaiting_image",
                    "transaction_type": transaction_type,
                    "category": text,
                },
            )
            return ask_for_receipt_after_category(text)
        if state.get("mode") == "awaiting_account_other_category":
            category = f"อื่นๆ - {text}" if text else "อื่นๆ"
            set_user_state(
                line_user_id,
                {
                    "mode": "awaiting_image",
                    "transaction_type": state.get("transaction_type") or "Expense",
                    "category": category,
                },
            )
            return ask_for_receipt_after_category(category)
        if state.get("mode") == "awaiting_hr_medical_certificate" and text not in {"เน€เธกเธเธน", "menu", "Menu", "MENU", "เธขเธเน€เธฅเธดเธ"}:
            return "เธเธฃเธธเธ“เธฒเธชเนเธเธฃเธนเธเนเธเธฃเธฑเธเธฃเธญเธเนเธเธ—เธขเนเธชเธณเธซเธฃเธฑเธเธงเธฑเธเธ—เธตเนเธฅเธฒเธเนเธงเธขเธเนเธฐ เธซเธฃเธทเธญเธเธดเธกเธเน เธขเธเน€เธฅเธดเธ เน€เธเธทเนเธญเน€เธฃเธดเนเธกเนเธซเธกเน"
        if state.get("mode") == "awaiting_hr_medical_certificate" and text == "เธขเธเน€เธฅเธดเธ":
            clear_user_state(line_user_id)
            return hr_menu_message()
        if text in {"เน€เธกเธเธน", "menu", "Menu", "MENU", "เธเธฑเธเธเธต"}:
            clear_user_state(line_user_id)

            return menu_message()
        approval_match = re.match(r"^HR_(APPROVE|REJECT):(.+)$", text, flags=re.IGNORECASE)
        if approval_match:
            approved = approval_match.group(1).upper() == "APPROVE"
            request_id = approval_match.group(2).strip()
            try:
                result = update_hr_approval_to_google(request_id, approved, line_user_id)
            except Exception as exc:
                runtime_log(f"Update HR approval failed: {exc}")
                return abort_flow_message(f"\u0e2d\u0e31\u0e1b\u0e40\u0e14\u0e15\u0e1c\u0e25\u0e2d\u0e19\u0e38\u0e21\u0e31\u0e15\u0e34\u0e44\u0e21\u0e48\u0e2a\u0e33\u0e40\u0e23\u0e47\u0e08\u0e04\u0e48\u0e30 ({exc})")
            status = result.get("approvalStatus") or ("\u0e2d\u0e19\u0e38\u0e21\u0e31\u0e15\u0e34" if approved else "\u0e44\u0e21\u0e48\u0e2d\u0e19\u0e38\u0e21\u0e31\u0e15\u0e34")
            requester_id = str(result.get("lineUserId") or "")
            request_type = result.get("requestType") or "\u0e04\u0e33\u0e02\u0e2d HR"
            employee_name = result.get("employeeName") or "-"
            if requester_id:
                push_line_messages(
                    requester_id,
                    [
                        text_message(
                            f"\u0e1c\u0e25\u0e01\u0e32\u0e23\u0e2d\u0e19\u0e38\u0e21\u0e31\u0e15\u0e34 {request_type}\n"
                            f"Request ID: {request_id}\n"
                            f"\u0e0a\u0e37\u0e48\u0e2d\u0e1e\u0e19\u0e31\u0e01\u0e07\u0e32\u0e19: {employee_name}\n"
                            f"\u0e2a\u0e16\u0e32\u0e19\u0e30: {status}"
                        )
                    ],
                )
            return (
                f"\u0e1a\u0e31\u0e19\u0e17\u0e36\u0e01\u0e1c\u0e25\u0e2d\u0e19\u0e38\u0e21\u0e31\u0e15\u0e34\u0e40\u0e23\u0e35\u0e22\u0e1a\u0e23\u0e49\u0e2d\u0e22\u0e04\u0e48\u0e30\n"
                f"Request ID: {request_id}\n"
                f"\u0e0a\u0e37\u0e48\u0e2d\u0e1e\u0e19\u0e31\u0e01\u0e07\u0e32\u0e19: {employee_name}\n"
                f"\u0e2a\u0e16\u0e32\u0e19\u0e30: {status}"
            )
        if text in {"เธชเธ•เนเธญเธ", "เธชเธ•เนเธญเธ", "stock", "Stock", "STOCK"}:
            clear_user_state(line_user_id)
            return stock_menu_message()
        if text in {"เธเธณเน€เธเนเธฒเธชเธ•เนเธญเธ", "เธเธณเธญเธญเธเธชเธ•เนเธญเธ", "เน€เธเนเธเธชเธ•เนเธญเธ"}:
            clear_user_state(line_user_id)
            return buttons_template_message(
                f"เน€เธกเธเธน {text}\n\n"
                "เธฃเธฐเธเธเธชเนเธงเธเธเธตเนเธเธณเธฅเธฑเธเน€เธ•เธฃเธตเธขเธกเนเธเนเธเธฒเธเธเนเธฐ",
                [
                    ("เธเธฅเธฑเธเน€เธกเธเธนเธชเธ•เนเธญเธ", "เธชเธ•เนเธญเธ"),
                    ("เธเธฅเธฑเธเน€เธกเธเธนเธเธฑเธเธเธต", "เธเธฑเธเธเธต"),
                ],
            )
        if text == "เธเนเธเธซเธฒเธชเธดเธเธเนเธฒ":
            clear_user_state(line_user_id)
            return stock_branch_menu_message()
        stock_branch_match = re.match("^(?:\u0e04\u0e49\u0e19\u0e2b\u0e32\u0e2a\u0e15\u0e47\u0e2d\u0e04|\u0e04\u0e49\u0e19\u0e2b\u0e32\u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32):(.+)$", text)
        if stock_branch_match:
            branch = stock_branch_match.group(1).strip()
            set_user_state(line_user_id, {"mode": "awaiting_stock_product_query", "stock_branch": branch})
            return (
                f"เน€เธฅเธทเธญเธเธชเธฒเธเธฒ {branch} เนเธฅเนเธงเธเนเธฐ\n"
                "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเธเธทเนเธญเธชเธดเธเธเนเธฒ/เธเธฒเธฃเนเนเธเนเธ” เธซเธฃเธทเธญเธชเนเธเธเธเธฒเธฃเนเนเธเนเธ”เนเธ”เนเน€เธฅเธข\n"
                "เธชเนเธเนเธ”เนเธชเธนเธเธชเธธเธ” 10 เธฃเธฒเธขเธเธฒเธฃเธ•เนเธญเธเธฃเธฑเนเธ เนเธ”เธขเธเธถเนเธเธเธฃเธฃเธ—เธฑเธ”เนเธซเธกเนเธซเธฃเธทเธญเธเธฑเนเธเธ”เนเธงเธข comma\n"
                "เธ•เธฑเธงเธญเธขเนเธฒเธ: 8851234567890, เธ•เธธเนเธเธ•เธฒ, ABC001"
            )
        if state.get("mode") == "awaiting_stock_product_query":
            branch = str(state.get("stock_branch") or "-")
            queries, total_entered = parse_stock_queries(text, 10)
            if not queries:
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเธเธทเนเธญเธชเธดเธเธเนเธฒเธซเธฃเธทเธญเธเธฒเธฃเนเนเธเนเธ”เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธ•เธฃเธงเธเธชเธญเธเธชเธ•เนเธญเธเธเนเธฐ"
            results: dict[str, list[dict[str, Any]]] = {}
            try:
                for query in queries:
                    results[query] = search_stock_product(branch, query)
            except Exception as exc:
                runtime_log(f"Stock search failed: {exc}")
                clear_user_state(line_user_id)
                return abort_flow_message(f"เธเนเธเธซเธฒเธชเธ•เนเธญเธเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
            set_user_state(line_user_id, {"mode": "awaiting_stock_product_query", "stock_branch": branch})
            return [
                text_message(format_stock_check_results(branch, queries, results, total_entered)),
                buttons_template_message(
                    "เธ•เนเธญเธเธเธฒเธฃเธ•เธฃเธงเธเธชเธญเธเธ•เนเธญเนเธซเธกเธเธฐ",
                    [
                        ("เธ•เธฃเธงเธเธชเธฒเธเธฒเน€เธ”เธดเธก", f"เธเนเธเธซเธฒเธชเธ•เนเธญเธ:{branch}"),
                        ("เน€เธฅเธทเธญเธเธชเธฒเธเธฒเนเธซเธกเน", "เธชเธ•เนเธญเธ"),
                    ],
                ),
            ]
        if text in {"HR", "hr", "Hr", "เธเนเธฒเธขเธเธธเธเธเธฅ"}:
            clear_user_state(line_user_id)
            return hr_menu_message()
        if text in {"เธชเธดเธเธเนเธฒ", "product", "Product", "PRODUCT"}:
            clear_user_state(line_user_id)
            return stock_branch_menu_message()
        if text == "เธ•เธฒเธฃเธฒเธเธเธฒเธ":
            clear_user_state(line_user_id)
            return schedule_month_menu_message()
        schedule_match = re.match(r"^เธ•เธฒเธฃเธฒเธเธเธฒเธ:([+-]?\d+)$", text)
        if schedule_match:
            clear_user_state(line_user_id)
            month_offset = max(-1, min(1, int(schedule_match.group(1))))
            try:
                result = get_hr_schedule_link(month_offset)
            except Exception as exc:
                runtime_log(f"Get HR schedule failed: {exc}")
                return abort_flow_message(f"เน€เธเธดเธ”เธ•เธฒเธฃเธฒเธเธเธฒเธเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
            url = result.get("url") or result.get("spreadsheetUrl") or ""
            pdf_url = result.get("pdfUrl") or result.get("pdf_url") or ""
            pdf_download_url = result.get("pdfDownloadUrl") or result.get("pdf_download_url") or pdf_url
            sheet_name = result.get("sheetName") or "HR_Work_Schedule"
            if pdf_download_url:
                try:
                    pdf_bytes = download_binary(str(pdf_download_url))
                    image_path = render_pdf_first_page_to_jpeg(pdf_bytes, f"hr_schedule_{month_offset}")
                    return [
                        text_message(f"เธ•เธฒเธฃเธฒเธเธเธฒเธ\n{sheet_name}"),
                        image_message(public_file_url(public_base_url, image_path)),
                    ]
                except Exception as exc:
                    runtime_log(f"Schedule PDF to JPEG failed: {exc}")
                return (
                    "เธ•เธฒเธฃเธฒเธเธเธฒเธ\n"
                    f"{sheet_name}\n"
                    f"{pdf_url or url}\n\n"
                    "เธฃเธฐเธเธเนเธเธฅเธเน€เธเนเธเธฃเธนเธเนเธกเนเธชเธณเน€เธฃเนเธเธเธฑเนเธงเธเธฃเธฒเธง เธเธถเธเธชเนเธเธฅเธดเธเธเนเธชเธณเธฃเธญเธเนเธซเนเธเนเธฐ\n"
                    f"เธฅเธดเธเธเน Google Sheet: {url}"
                )
            return f"เธ•เธฒเธฃเธฒเธเธเธฒเธ\n{sheet_name}\n{url}"
        if text in {"เธฅเธฒเธเนเธงเธข", "เธฅเธฒเธเธดเธ", "เนเธเนเธเธเธญเธงเธฑเธเธซเธขเธธเธ”เธฅเนเธงเธเธซเธเนเธฒ", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเน€เธงเธฅเธฒเน€เธเนเธฒ-เธญเธญเธเธเธฒเธ", "เนเธเนเธเน€เธเธฅเธตเนเธขเธเธงเธฑเธเธ—เธณเธเธฒเธ"}:
            draft = hr_request_blank(text, line_user_id)
            set_user_state(
                line_user_id,
                {
                    "mode": "awaiting_hr_form",
                    "hr_request": draft,
                },
            )
            return hr_request_form(draft)
        if state.get("mode") == "awaiting_substitute_receipt_decision":
            if text == "1":
                match = state.get("substitute_match")
                if not isinstance(match, dict) or not can_create_substitute_receipt(match):
                    clear_user_state(line_user_id)
                    return "เนเธกเนเธเธเธเนเธญเธกเธนเธฅเธชเธณเธซเธฃเธฑเธเธชเธฃเนเธฒเธเนเธเนเธ—เธเธเนเธฐ เธเธฃเธธเธ“เธฒเน€เธฃเธดเนเธกเธฃเธฒเธขเธเธฒเธฃเนเธซเธกเนเธญเธตเธเธเธฃเธฑเนเธ"
                clear_user_state(line_user_id)
                return substitute_receipt_messages([match], public_base_url, line_user_id)
            if text == "2":
                clear_user_state(line_user_id)
                return [
                    text_message("เธฃเธฑเธเธ—เธฃเธฒเธเธเนเธฐ เนเธกเนเธชเธฃเนเธฒเธเนเธเนเธ—เธเธชเธณเธซเธฃเธฑเธเธฃเธฒเธขเธเธฒเธฃเธเธตเน\nเธชเธฒเธกเธฒเธฃเธ–เน€เธฃเธดเนเธกเธ—เธณเธฃเธฒเธขเธเธฒเธฃเนเธซเธกเนเนเธ”เนเน€เธฅเธขเธเนเธฐ"),
                    menu_message(),
                ]
            return quick_reply_text_message(
                "เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเธงเนเธฒเธ•เนเธญเธเธเธฒเธฃเธชเธฃเนเธฒเธเนเธเนเธ—เธเธซเธฃเธทเธญเนเธกเนเธเธฐ\n\n"
                "1 = เธ•เนเธญเธเธเธฒเธฃเธชเธฃเนเธฒเธเนเธเนเธ—เธ\n"
                "2 = เนเธกเนเธ•เนเธญเธเธเธฒเธฃ",
                [
                    ("๐งพ 1 เธชเธฃเนเธฒเธเนเธเนเธ—เธ", "1"),
                    ("เนเธกเนเธชเธฃเนเธฒเธเนเธเนเธ—เธ", "2"),
                ],
            )
        if state.get("mode") == "awaiting_substitute_select":
            row_match = re.match(r"^(?:row\s*)?(\d+)$", text, flags=re.IGNORECASE)
            if not row_match:
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเน€เธฅเธ Row เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธชเธฃเนเธฒเธเนเธเนเธ—เธ เน€เธเนเธ Row 12"
            wanted_row = int(row_match.group(1))
            matches = [item for item in state.get("substitute_matches", []) if int(item.get("row", 0)) == wanted_row]
            if not matches:
                return "เนเธกเนเธเธ Row เธเธตเนเธเธฒเธเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธเนเธเธซเธฒ เธเธฃเธธเธ“เธฒเธเธดเธกเธเน Row เนเธซเธกเนเธเนเธฐ"
            clear_user_state(line_user_id)
            return substitute_receipt_messages(matches, public_base_url, line_user_id)
        substitute_match = re.match(r"^เนเธเนเธ—เธ\s+(.+)$", text, flags=re.IGNORECASE)
        if substitute_match:
            amount = normalize_amount(substitute_match.group(1))
            if amount is None:
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเธเธณเธชเธฑเนเธ เน€เธเนเธ เนเธเนเธ—เธ 59 เธซเธฃเธทเธญ เนเธเนเธ—เธ 12705.18"
            try:
                matches = search_google_sheet_by_total(amount)
            except Exception as exc:
                runtime_log(f"Substitute receipt search failed: {exc}")
                clear_user_state(line_user_id)
                return abort_flow_message(f"เธเนเธเธซเธฒเธฃเธฒเธขเธเธฒเธฃเน€เธเธทเนเธญเธชเธฃเนเธฒเธเนเธเนเธ—เธเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
            matches = [item for item in matches if can_create_substitute_receipt(item)]
            if not matches:
                return (
                    f"เนเธกเนเธเธเธฃเธฒเธขเธเธฒเธฃเธเนเธฒเนเธเนเธเนเธฒเธขเธขเธญเธ” {amount:,.2f} เธ—เธตเนเธชเธฒเธกเธฒเธฃเธ–เธชเธฃเนเธฒเธเนเธเนเธ—เธเนเธ”เนเธเนเธฐ\n"
                    "เธ•เธฃเธงเธเธชเธญเธเธงเนเธฒเธขเธญเธ”เธ•เธฃเธเธเธฑเธ Google Sheet เธซเธฃเธทเธญเธเธดเธกเธเนเธขเธญเธ”เธฃเธงเธกเธชเธธเธ—เธเธดเธเธญเธเธฃเธฒเธขเธเธฒเธฃเธเธฑเนเธเธญเธตเธเธเธฃเธฑเนเธ"
                )
            if len(matches) == 1:
                return substitute_receipt_messages(matches, public_base_url, line_user_id)
            set_user_state(line_user_id, {"mode": "awaiting_substitute_select", "substitute_matches": matches[:10]})
            return format_google_sheet_matches(matches, "เธเธเธซเธฅเธฒเธขเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธชเธฒเธกเธฒเธฃเธ–เธชเธฃเนเธฒเธเนเธเนเธ—เธเนเธ”เน") + "\n\nเธเธฃเธธเธ“เธฒเธเธดเธกเธเนเน€เธฅเธ Row เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธชเธฃเนเธฒเธเนเธเนเธ—เธ"
        if state.get("mode") == "awaiting_confirmation" and text in {"1", "เธ•เธฃเธงเธเธชเธญเธเนเธฅเธฐเธขเธทเธเธขเธฑเธ"}:
            if not state.get("pending_data"):
                return "เธขเธฑเธเนเธกเนเธกเธตเธเธดเธฅเธ—เธตเนเธฃเธญเธขเธทเธเธขเธฑเธเธเนเธฐ เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเน€เธกเธเธน เธเธดเธฅเธฃเธฒเธขเธฃเธฑเธ เธซเธฃเธทเธญ เธเธดเธฅเธฃเธฒเธขเธเนเธฒเธข เธเนเธญเธ"
            return confirm_pending_to_google(line_user_id, state, public_base_url)
        if state.get("mode") == "awaiting_submitter_name":
            pending = deserialize_data(state.get("pending_data", {}))
            pending["submitter_name"] = text
            state["mode"] = "awaiting_confirmation"
            state["pending_data"] = serialize_data(pending)
            set_user_state(line_user_id, state)
            return confirmation_prompt(pending)
        if state.get("mode") == "awaiting_cancel_total":
            amount = normalize_amount(text)
            if amount is None:
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเธขเธญเธ”เธฃเธงเธกเธชเธธเธ—เธเธดเธซเธฃเธทเธญเธขเธญเธ”เธเนเธญเธ VAT เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธขเธเน€เธฅเธดเธ เน€เธเนเธ 2251.72"
            try:
                matches = search_google_sheet_by_total(amount)
            except Exception as exc:
                runtime_log(f"Cancel search failed: {exc}")
                clear_user_state(line_user_id)
                return abort_flow_message(f"เธเนเธเธซเธฒเธฃเธฒเธขเธเธฒเธฃเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
            if not matches:
                clear_user_state(line_user_id)
                return f"เนเธกเนเธเธเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธกเธตเธขเธญเธ” {amount:,.2f} เธเนเธฐ เธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเธชเธญเธเธงเนเธฒเธขเธญเธ”เธเธตเนเธ•เธฃเธเธเธฑเธเธขเธญเธ”เธฃเธงเธกเธชเธธเธ—เธเธดเธซเธฃเธทเธญเธขเธญเธ”เธเนเธญเธ VAT เนเธ Google Sheet"
            if len(matches) == 1:
                state["mode"] = "awaiting_cancel_confirm"
                state["cancel_row"] = int(matches[0]["row"])
                state["cancel_sheet"] = str(matches[0].get("sheetName") or "")
                set_user_state(line_user_id, state)
                return format_google_sheet_matches(matches, "เธเธเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธขเธเน€เธฅเธดเธ") + "\n\nเธ•เธญเธ 1 = เธขเธทเธเธขเธฑเธเธขเธเน€เธฅเธดเธ\nเธ•เธญเธ 2 = เนเธกเนเธขเธเน€เธฅเธดเธ"
            state["mode"] = "awaiting_cancel_select"
            state["cancel_matches"] = matches[:10]
            set_user_state(line_user_id, state)
            return format_google_sheet_matches(matches, "เธเธเธซเธฅเธฒเธขเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธกเธตเธขเธญเธ”เธฃเธงเธกเธเธตเน") + "\n\nเธเธฃเธธเธ“เธฒเธเธดเธกเธเนเน€เธฅเธ Row เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธขเธเน€เธฅเธดเธ"
        if state.get("mode") == "awaiting_cancel_select":
            row_match = re.match(r"^(?:row\s*)?(\d+)$", text, flags=re.IGNORECASE)
            if not row_match:
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเน€เธฅเธ Row เธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธขเธเน€เธฅเธดเธ เน€เธเนเธ Row 12"
            wanted_row = int(row_match.group(1))
            matches = [item for item in state.get("cancel_matches", []) if int(item.get("row", 0)) == wanted_row]
            if not matches:
                return "เนเธกเนเธเธ Row เธเธตเนเธเธฒเธเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธเนเธเธซเธฒ เธเธฃเธธเธ“เธฒเธเธดเธกเธเน Row เนเธซเธกเนเธเนเธฐ"
            state["mode"] = "awaiting_cancel_confirm"
            state["cancel_row"] = wanted_row
            state["cancel_sheet"] = str(matches[0].get("sheetName") or "")
            set_user_state(line_user_id, state)
            return format_google_sheet_matches(matches, "เธขเธทเธเธขเธฑเธเธฃเธฒเธขเธเธฒเธฃเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเธขเธเน€เธฅเธดเธ") + "\n\nเธ•เธญเธ 1 = เธขเธทเธเธขเธฑเธเธขเธเน€เธฅเธดเธ\nเธ•เธญเธ 2 = เนเธกเนเธขเธเน€เธฅเธดเธ"
        if state.get("mode") == "awaiting_cancel_confirm":
            if text == "1":
                row = int(state.get("cancel_row", 0))
                try:
                    delete_google_sheet_row(row, str(state.get("cancel_sheet") or ""))
                except Exception as exc:
                    runtime_log(f"Cancel delete failed: {exc}")
                    clear_user_state(line_user_id)
                    return abort_flow_message(f"เธขเธเน€เธฅเธดเธเธฃเธฒเธขเธเธฒเธฃเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
                clear_user_state(line_user_id)
                return "เธขเธเน€เธฅเธดเธเธเธดเธฅเน€เธฃเธตเธขเธเธฃเนเธญเธข"
            if text == "2":
                clear_user_state(line_user_id)
                return "เธขเธเน€เธฅเธดเธเธเธณเธชเธฑเนเธเนเธฅเนเธงเธเนเธฐ"
            return "เธเธฃเธธเธ“เธฒเธ•เธญเธ 1 เน€เธเธทเนเธญเธขเธทเธเธขเธฑเธเธขเธเน€เธฅเธดเธ เธซเธฃเธทเธญ 2 เน€เธเธทเนเธญเนเธกเนเธขเธเน€เธฅเธดเธ"
        if state.get("mode") == "awaiting_duplicate_confirmation":
            if text == "1":
                state["mode"] = "awaiting_duplicate_edit_choice"
                set_user_state(line_user_id, state)
                return "เธ•เนเธญเธเธเธฒเธฃเนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเธเธญเธเธฃเธฒเธขเธเธฒเธฃเน€เธ”เธดเธกเธซเธฃเธทเธญเนเธกเน?\nเธ•เธญเธ 1 = เนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ\nเธ•เธญเธ 2 = เนเธกเนเนเธเนเนเธเนเธฅเธฐเนเธกเนเธเธฑเธเธ—เธถเธเธเนเธณ"
            if text == "2":
                state["mode"] = "awaiting_confirmation"
                state["duplicate_checked"] = True
                set_user_state(line_user_id, state)
                return confirm_pending_to_google(line_user_id, state, public_base_url)
            return "เธเธฃเธธเธ“เธฒเธ•เธญเธ 1 = เน€เธเนเธเธฃเธฒเธขเธเธฒเธฃเน€เธ”เธตเธขเธงเธเธฑเธ เธซเธฃเธทเธญ 2 = เธเธฑเธเธ—เธถเธเน€เธเนเธเธฃเธฒเธขเธเธฒเธฃเนเธซเธกเน"
        if state.get("mode") == "awaiting_duplicate_edit_choice":
            if text == "1":
                state["mode"] = "awaiting_duplicate_doc_type"
                set_user_state(line_user_id, state)
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเนเธซเธกเนเธ—เธตเนเธ•เนเธญเธเธเธฒเธฃเนเธเนเนเธ เน€เธเนเธ เนเธเธเธณเธเธฑเธเธ เธฒเธฉเธต / เนเธเน€เธชเธฃเนเธ / เธเธดเธฅ"
            if text == "2":
                clear_user_state(line_user_id)
                return "เธขเธเน€เธฅเธดเธเธเธฒเธฃเธเธฑเธเธ—เธถเธเธเนเธณเนเธฅเนเธงเธเนเธฐ"
            return "เธเธฃเธธเธ“เธฒเธ•เธญเธ 1 เน€เธเธทเนเธญเนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃ เธซเธฃเธทเธญ 2 เน€เธเธทเนเธญเนเธกเนเนเธเนเนเธ"
        if state.get("mode") == "awaiting_duplicate_doc_type":
            state["new_document_type"] = text
            state["mode"] = "awaiting_duplicate_update_confirm"
            set_user_state(line_user_id, state)
            matches = state.get("duplicate_matches", [])
            return format_google_sheet_matches(matches, "เธฃเธฒเธขเธเธฒเธฃเน€เธ”เธดเธกเธ—เธตเนเธเธฐเธ–เธนเธเนเธเนเนเธ") + f"\n\nเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเนเธซเธกเน: {text}\nเธ•เธญเธ 1 = เธขเธทเธเธขเธฑเธเนเธเนเนเธ\nเธ•เธญเธ 2 = เนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเธญเธตเธเธเธฃเธฑเนเธ"
        if state.get("mode") == "awaiting_duplicate_update_confirm":
            if text == "1":
                matches = state.get("duplicate_matches", [])
                if not matches:
                    clear_user_state(line_user_id)
                    return "เนเธกเนเธเธเธฃเธฒเธขเธเธฒเธฃเน€เธ”เธดเธกเธชเธณเธซเธฃเธฑเธเนเธเนเนเธเธเนเธฐ"
                target = matches[0]
                row = int(target["row"])
                sheet_name = str(target.get("sheetName") or "")
                document_type = str(state.get("new_document_type") or "")
                try:
                    update_google_sheet_document_type(row, document_type, sheet_name)
                except Exception as exc:
                    runtime_log(f"Duplicate doc type update failed: {exc}")
                    clear_user_state(line_user_id)
                    return abort_flow_message(f"เนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
                clear_user_state(line_user_id)
                return "เนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเน€เธฃเธตเธขเธเธฃเนเธญเธข"
            if text == "2":
                state["mode"] = "awaiting_duplicate_doc_type"
                set_user_state(line_user_id, state)
                return "เธเธฃเธธเธ“เธฒเธเธดเธกเธเนเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเนเธซเธกเนเธญเธตเธเธเธฃเธฑเนเธเธเนเธฐ"
            return "เธเธฃเธธเธ“เธฒเธ•เธญเธ 1 เน€เธเธทเนเธญเธขเธทเธเธขเธฑเธเนเธเนเนเธ เธซเธฃเธทเธญ 2 เน€เธเธทเนเธญเนเธเนเนเธเธเธฃเธฐเน€เธ เธ—เน€เธญเธเธชเธฒเธฃเธญเธตเธเธเธฃเธฑเนเธ"
        if state.get("mode") == "awaiting_confirmation" and text in {"2", "เนเธเนเนเธ"}:
            if not state.get("pending_data"):
                return "เธขเธฑเธเนเธกเนเธกเธตเธเธดเธฅเธ—เธตเนเธฃเธญเนเธเนเนเธเธเนเธฐ"
            state["mode"] = "awaiting_correction"
            set_user_state(line_user_id, state)
            return manual_entry_form(deserialize_data(state.get("pending_data", {})))
        if state.get("mode") == "awaiting_hr_form":
            draft = dict(state.get("hr_request") or {})
            updated = parse_hr_request_text(text, draft)
            state["mode"] = "awaiting_hr_confirm"
            state["hr_request"] = updated
            set_user_state(line_user_id, state)
            return hr_confirm_message(updated)
        if state.get("mode") == "awaiting_hr_confirm":
            if text == "1":
                request_data = dict(state.get("hr_request") or {})
                if not str(request_data.get("employee_name") or "").strip():
                    state["mode"] = "awaiting_hr_form"
                    set_user_state(line_user_id, state)
                    return "เธเธฃเธธเธ“เธฒเธฃเธฐเธเธธเธเธทเนเธญเธเธเธฑเธเธเธฒเธเธเนเธญเธเธชเนเธเธเธณเธเธญเธเนเธฐ\n\n" + hr_request_form(request_data)
                try:
                    result = save_hr_request_to_google(request_data, line_user_id)
                except Exception as exc:
                    runtime_log(f"Save HR request failed: {exc}")
                    clear_user_state(line_user_id)
                    return abort_flow_message(f"เธเธฑเธเธ—เธถเธเธเธณเธเธญ HR เนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
                request_id = str(result.get("requestId") or result.get("row") or uuid.uuid4().hex[:8])
                request_data["request_id"] = request_id
                approver_id = str(CONFIG.get("hr_approver_line_id") or os.getenv("HR_APPROVER_LINE_ID") or "Ud260925c43fb0823fea42224a2929393")
                push_line_messages(
                    approver_id,
                    [
                        text_message(
                            "เธกเธตเธเธณเธเธญ HR เธฃเธญเธญเธเธธเธกเธฑเธ•เธด\n\n"
                            + format_hr_request(request_data)
                            + f"\n\nRequest ID: {request_id}"
                        ),
                        approval_buttons_message(request_id),
                    ],
                )
                if request_data.get("request_type") == "เธฅเธฒเธเนเธงเธข":
                    set_user_state(
                        line_user_id,
                        {
                            "mode": "awaiting_hr_medical_certificate",
                            "hr_request": request_data,
                            "hr_request_id": request_id,
                        },
                    )
                    return (
                        "เธชเนเธเธเธณเธเธญเธฅเธฒเธเนเธงเธขเน€เธฃเธตเธขเธเธฃเนเธญเธขเธเนเธฐ\n"
                        f"Request ID: {request_id}\n\n"
                        "เธเธฃเธธเธ“เธฒเธชเนเธเธฃเธนเธเนเธเธฃเธฑเธเธฃเธญเธเนเธเธ—เธขเนเธชเธณเธซเธฃเธฑเธเธงเธฑเธเธ—เธตเนเธฅเธฒเธเนเธงเธขเน€เธเนเธฒเธกเธฒเนเธ”เนเน€เธฅเธขเธเนเธฐ"
                    )
                clear_user_state(line_user_id)
                return f"เธชเนเธเธเธณเธเธญ HR เน€เธฃเธตเธขเธเธฃเนเธญเธขเธเนเธฐ\nRequest ID: {request_id}\nเธชเธ–เธฒเธเธฐ: เธฃเธญเธญเธเธธเธกเธฑเธ•เธด"
            if text == "2":
                state["mode"] = "awaiting_hr_form"
                set_user_state(line_user_id, state)
                return hr_request_form(dict(state.get("hr_request") or {}))
            return "เธเธฃเธธเธ“เธฒเธ•เธญเธ 1 เน€เธเธทเนเธญเธขเธทเธเธขเธฑเธเธชเนเธเธเธณเธเธญ เธซเธฃเธทเธญ 2 เน€เธเธทเนเธญเนเธเนเนเธเธเนเธญเธกเธนเธฅเธเนเธฐ"
        if text in {"1", "บิลรายรับ", "เธเธดเธฅเธฃเธฒเธขเธฃเธฑเธ"}:
            set_user_state(line_user_id, {"mode": "awaiting_account_category", "transaction_type": "Revenue"})
            return category_menu_message("Revenue")
        if text in {"2", "บิลรายจ่าย", "เธเธดเธฅเธฃเธฒเธขเธเนเธฒเธข"}:
            set_user_state(line_user_id, {"mode": "awaiting_account_category", "transaction_type": "Expense"})
            return category_menu_message("Expense")
        if text in {"3", "เรียกดูบัญชี", "เรียกดูรายละเอียดบัญชี", "เน€เธฃเธตเธขเธเธ”เธนเธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธเธฑเธเธเธต"}:
            set_user_state(line_user_id, {"mode": "awaiting_lookup_bill_no"})
            return "กรุณาพิมพ์เลขที่บิล ชื่อร้าน/คู่ค้า หรือยอดรวมที่ต้องการตรวจสอบค่ะ"
        if text in {"4", "ยกเลิกรายการ", "ยกเลิกการทำรายการ", "เธขเธเน€เธฅเธดเธเธเธฒเธฃเธ—เธณเธฃเธฒเธขเธเธฒเธฃ"}:
            set_user_state(line_user_id, {"mode": "awaiting_cancel_total"})
            return "กรุณาพิมพ์ยอดรวมสุทธิหรือยอดก่อน VAT ของรายการที่ต้องการยกเลิกค่ะ เช่น 2251.72"
        if state.get("mode") == "awaiting_lookup_bill_no":
            row_match = re.match(r"^(?:row\s*)?(\d+)$", text, flags=re.IGNORECASE)
            if row_match and state.get("lookup_rows"):
                wanted_row = int(row_match.group(1))
                matched_rows = [item for item in state.get("lookup_rows", []) if int(item.get("row", 0)) == wanted_row]
                if not matched_rows:
                    return "เนเธกเนเธเธ Row เธ—เธตเนเน€เธฅเธทเธญเธเธเธฒเธเธเธฅเธเธฒเธฃเธเนเธเธซเธฒเธเนเธญเธเธซเธเนเธฒ เธเธฃเธธเธ“เธฒเธเธดเธกเธเน Row เนเธซเธกเนเธเนเธฐ"
                found = find_transaction_by_row(wanted_row)
                clear_user_state(line_user_id)
                if not found:
                    return f"เนเธกเนเธเธ Row {wanted_row} เนเธ Excel เธเนเธฐ"
                sheet_name, row, data = found
                image_path = render_row_summary_image(sheet_name, row, data, "เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธเธฑเธเธเธต")
                return [
                    text_message(f"เธเธเธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธ—เธตเน {sheet_name}!Row {row}"),
                    image_message(public_file_url(public_base_url, image_path)),
                ]

            results = search_transactions(text, max_results=5)
            if not results:
                clear_user_state(line_user_id)
                return f"เนเธกเนเธเธเธเนเธญเธกเธนเธฅเธเธฒเธเธเธณเธเนเธ: {text}\nเธเธฃเธธเธ“เธฒเธ•เธฃเธงเธเน€เธฅเธเธ—เธตเนเธเธดเธฅ เธเธทเนเธญเธฃเนเธฒเธ เธซเธฃเธทเธญเธขเธญเธ”เธฃเธงเธกเธญเธตเธเธเธฃเธฑเนเธเธเนเธฐ"
            if len(results) > 1:
                set_user_state(
                    line_user_id,
                    {
                        "mode": "awaiting_lookup_bill_no",
                        "lookup_rows": [{"sheet": sheet_name, "row": row} for sheet_name, row, _ in results],
                    },
                )
                return format_lookup_results(results, text)

            clear_user_state(line_user_id)
            sheet_name, row, data = results[0]
            image_path = render_row_summary_image(sheet_name, row, data, "เธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธเธฑเธเธเธต")
            return [
                text_message(f"เธเธเธฃเธฒเธขเธฅเธฐเน€เธญเธตเธขเธ”เธเธฒเธเธเธณเธเนเธ {text} เธ—เธตเน {sheet_name}!Row {row}"),
                image_message(public_file_url(public_base_url, image_path)),
            ]
        if text == "เธ•เธฃเธงเธเธชเธญเธเนเธฅเธฐเธขเธทเธเธขเธฑเธ":
            if state.get("mode") != "awaiting_confirmation" or not state.get("pending_data"):
                return "เธขเธฑเธเนเธกเนเธกเธตเธเธดเธฅเธ—เธตเนเธฃเธญเธขเธทเธเธขเธฑเธเธเนเธฐ เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเน€เธกเธเธน เธเธดเธฅเธฃเธฒเธขเธฃเธฑเธ เธซเธฃเธทเธญ เธเธดเธฅเธฃเธฒเธขเธเนเธฒเธข เธเนเธญเธ"
            pending = deserialize_data(state["pending_data"])
            image_path = Path(state.get("image_path", ""))
            if not image_path.exists():
                return "เนเธกเนเธเธเนเธเธฅเนเธฃเธนเธเน€เธญเธเธชเธฒเธฃเน€เธ”เธดเธกเธเนเธฐ เธเธฃเธธเธ“เธฒเธชเนเธเน€เธญเธเธชเธฒเธฃเนเธซเธกเนเธญเธตเธเธเธฃเธฑเนเธ"
            sheet_name = "Google Sheet"
            row = "-"
            result = None
            try:
                result = send_to_google_sheet(pending, image_path, line_user_id, public_base_url)
            except Exception as exc:
                runtime_log(f"Google Sheet save failed: {exc}")
                clear_user_state(line_user_id)
                return abort_flow_message(f"Google Sheet: เธขเธฑเธเนเธกเนเธชเธณเน€เธฃเนเธ ({exc})")
            notify_accounting_import_approver(pending, line_user_id)
            clear_user_state(line_user_id)
            summary_image = render_row_summary_image(sheet_name, row, pending, "เธเธดเธฅเธเธณเน€เธเนเธฒ")
            messages = [
                text_message("เธเธฑเธเธ—เธถเธเน€เธฃเธตเธขเธเธฃเนเธญเธข\nGoogle Sheet: เธเธฑเธเธ—เธถเธเธชเธณเน€เธฃเนเธ"),
                image_message(public_file_url(public_base_url, summary_image)),
            ]
            substitute_match_data = substitute_match_from_pending(pending, result)
            if can_create_substitute_receipt(substitute_match_data):
                messages.extend(substitute_receipt_messages([substitute_match_data], public_base_url, line_user_id))
            runtime_log(
                f"Confirmed LINE receipt -> Google Sheet "
                f"type={pending.get('transaction_type')} date={pending['date']} total={pending['total']}"
            )
            return messages
        if text == "เนเธเนเนเธ":
            if state.get("mode") != "awaiting_confirmation" or not state.get("pending_data"):
                return "เธขเธฑเธเนเธกเนเธกเธตเธเธดเธฅเธ—เธตเนเธฃเธญเนเธเนเนเธเธเนเธฐ"
            state["mode"] = "awaiting_correction"
            set_user_state(line_user_id, state)
            return correction_form(deserialize_data(state.get("pending_data", {})))
        if state.get("mode") == "awaiting_correction":
            pending = deserialize_data(state.get("pending_data", {}))
            corrected = parse_correction_text_v2(text, pending)
            state["mode"] = "awaiting_confirmation"
            state["pending_data"] = serialize_data(corrected)
            set_user_state(line_user_id, state)
            return confirmation_prompt(corrected)
        if text.startswith("เนเธเนเนเธเธเธดเธฅ+") or text.startswith("เน€เธยเน€เธยเน€เธยเน€เธยเน€เธยเน€เธยเน€เธเธ”เน€เธเธ…+"):
            bill_no = text.split("+", 1)[1].strip()
            return delete_bill_from_excel(bill_no, line_user_id)
        delete_row_match = re.match(r"^delete\s+row\s+(\d+)$", text, flags=re.IGNORECASE)
        if delete_row_match:
            return delete_excel_row(delete_row_match.group(1), line_user_id)
        return menu_message()

    if message.get("type") not in {"image", "file"}:
        return menu_message() if not state.get("mode") else menu_text()

    if state.get("mode") == "awaiting_hr_medical_certificate":
        token = os.getenv(CONFIG["line"]["channel_access_token_env"])
        if not token:
            raise RuntimeError(f"Missing {CONFIG['line']['channel_access_token_env']}")
        try:
            runtime_log("Downloading LINE medical certificate content")
            image_path = download_line_content(message["id"], token, resolve_path(CONFIG["image_archive_dir"]))
            request_id = str(state.get("hr_request_id") or "")
            result = save_medical_certificate_to_google(request_id, image_path, line_user_id)
        except Exception as exc:
            runtime_log(f"Medical certificate save failed: {exc}")
            clear_user_state(line_user_id)
            return abort_flow_message(f"เธเธฑเธเธ—เธถเธเนเธเธฃเธฑเธเธฃเธญเธเนเธเธ—เธขเนเนเธกเนเธชเธณเน€เธฃเนเธเธเนเธฐ ({exc})")
        request_data = dict(state.get("hr_request") or {})
        approver_id = str(CONFIG.get("hr_approver_line_id") or os.getenv("HR_APPROVER_LINE_ID") or "Ud260925c43fb0823fea42224a2929393")
        file_url = result.get("fileUrl") or result.get("url") or ""
        push_line_messages(
            approver_id,
            [
                text_message(
                    "เนเธ”เนเธฃเธฑเธเนเธเธฃเธฑเธเธฃเธญเธเนเธเธ—เธขเนเธชเธณเธซเธฃเธฑเธเธเธณเธเธญเธฅเธฒเธเนเธงเธข\n\n"
                    + format_hr_request(request_data)
                    + f"\n\nRequest ID: {request_id}\nเนเธเธฅเน: {file_url}"
                ),
                approval_buttons_message(request_id),
            ],
        )
        clear_user_state(line_user_id)
        return f"เธเธฑเธเธ—เธถเธเนเธเธฃเธฑเธเธฃเธญเธเนเธเธ—เธขเนเน€เธฃเธตเธขเธเธฃเนเธญเธขเธเนเธฐ\nRequest ID: {request_id}\n{file_url}"

    if state.get("mode") not in {"awaiting_image", "awaiting_more_documents"}:
        return [
            text_message("เธเธฃเธธเธ“เธฒเน€เธฅเธทเธญเธเน€เธกเธเธนเธเนเธญเธเธชเนเธเธฃเธนเธเธเนเธฐ"),
            menu_message(),
        ]

    token = os.getenv(CONFIG["line"]["channel_access_token_env"])
    if not token:
        raise RuntimeError(f"Missing {CONFIG['line']['channel_access_token_env']}")

    try:
        runtime_log("Downloading LINE image content")
        image_path = download_line_content(message["id"], token, resolve_path(CONFIG["image_archive_dir"]))
        append_document_path_to_state(state, image_path)
        runtime_log(f"Downloaded LINE image to {image_path}")
        runtime_log("OCR started")
        text = ocr_image(image_path)
        runtime_log(f"OCR completed characters={len(text)}")
    except Exception as exc:
        runtime_log(f"OCR failed: {exc}")
        draft = blank_manual_entry(state.get("transaction_type", "Expense"))
        if state.get("category"):
            draft["category"] = state.get("category")
        draft = apply_revenue_payments_to_data(draft, state)
        if state.get("document_paths"):
            draft["document_paths"] = state.get("document_paths") or []
            draft["description"] = (draft.get("description") or "Manual entry from LINE") + f" | เอกสาร {len(state.get('document_paths') or [])} ชิ้น"
        draft["submitter_name"] = get_line_display_name(line_user_id)
        set_user_state(
            line_user_id,
            {
                "mode": "awaiting_correction",
                "transaction_type": draft["transaction_type"],
                "image_path": str(state.get("image_path") or image_path) if "image_path" in locals() else "",
                "document_paths": state.get("document_paths") or [],
                "pending_data": serialize_data(draft),
            },
        )
        return manual_entry_form(draft)
    parsed = parse_receipt_text(text, float(CONFIG.get("vat_rate", 0.07)))
    parsed = apply_transaction_type_defaults(parsed, state.get("transaction_type", "Expense"))
    if state.get("category"):
        parsed["category"] = state.get("category")
    parsed = apply_revenue_payments_to_data(parsed, state)
    if state.get("pending_data"):
        previous = deserialize_data(state.get("pending_data", {}))
        for key in ("transaction_type", "category", "before_vat", "vat", "total", "claimable", "revenue_branch", "revenue_payments"):
            if previous.get(key) not in (None, ""):
                parsed[key] = previous[key]
        parsed["raw_text"] = "\n\n--- เอกสารเพิ่มเติม ---\n\n".join(
            part for part in [str(previous.get("raw_text") or ""), str(parsed.get("raw_text") or "")] if part
        )
        parsed["description"] = previous.get("description") or parsed.get("description")
    document_count = len(state.get("document_paths") or [])
    if document_count:
        parsed["document_paths"] = state.get("document_paths") or []
        parsed["description"] = (parsed.get("description") or "") + f" | เอกสาร {document_count} ชิ้น"
    parsed["submitter_name"] = parsed.get("submitter_name") or get_line_display_name(line_user_id)
    set_user_state(
        line_user_id,
        {
            "mode": "awaiting_more_documents",
            "transaction_type": parsed["transaction_type"],
            "image_path": str(state.get("image_path") or image_path),
            "document_paths": state.get("document_paths") or [str(image_path)],
            "category": parsed.get("category") or state.get("category"),
            "revenue_branch": state.get("revenue_branch"),
            "revenue_payments": state.get("revenue_payments") or {},
            "pending_data": serialize_data(parsed),
        },
    )
    runtime_log(
        f"Parsed LINE receipt pending confirmation type={parsed['transaction_type']} "
        f"date={parsed['date']} total={parsed['total']} vendor={parsed['vendor']!r}"
    )
    return [
        text_message(
            "รับเอกสารแล้วค่ะ\n"
            f"เอกสารในรายการนี้: {len(state.get('document_paths') or [str(image_path)])} ชิ้น\n\n"
            "หากมีเอกสารของรายการเดียวกันเพิ่มเติม เช่น รายงานยอดขาย/สลิปโอนเงิน สามารถส่งต่อได้ค่ะ"
        ),
        continue_document_menu_message(),
    ]


class LineWebhookHandler(BaseHTTPRequestHandler):
    server_version = "LineExpenseBot/2.0"

    def _send_json(self, status: int, payload: dict[str, Any]) -> None:
        data = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        if self.path == "/health":
            self._send_json(HTTPStatus.OK, {"status": "ok", "version": self.server_version})
            return
        if self.path.startswith("/files/"):
            filename = Path(unquote(self.path.split("/files/", 1)[1])).name
            file_path = reply_image_dir() / filename
            if not file_path.exists() or not file_path.is_file():
                archive_path = resolve_path(CONFIG.get("image_archive_dir", "outputs/line_uploaded_receipts")) / filename
                if archive_path.exists() and archive_path.is_file():
                    file_path = archive_path
            if not file_path.exists() or not file_path.is_file():
                self._send_json(HTTPStatus.NOT_FOUND, {"error": "file not found"})
                return
            data = file_path.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", mimetypes.guess_type(str(file_path))[0] or "application/octet-stream")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self._send_json(HTTPStatus.NOT_FOUND, {"error": "not found"})

    def do_POST(self) -> None:
        try:
            runtime_log(f"POST received path={self.path}")
            if self.path != "/callback":
                self._send_json(HTTPStatus.NOT_FOUND, {"error": "not found"})
                return

            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            runtime_log(f"POST /callback content_length={length} signature_present={bool(self.headers.get('X-Line-Signature'))}")

            secret = os.getenv(CONFIG["line"]["channel_secret_env"])
            if not secret:
                runtime_log(f"Webhook rejected: missing {CONFIG['line']['channel_secret_env']} in this PowerShell session")
                self._send_json(HTTPStatus.OK, {"status": "accepted_with_error", "error": "missing channel secret"})
                return

            signature = self.headers.get("X-Line-Signature", "")
            if not verify_line_signature(body, signature, secret):
                runtime_log("Webhook rejected: invalid X-Line-Signature")
                self._send_json(HTTPStatus.OK, {"status": "accepted_with_error", "error": "invalid signature"})
                return

            payload = json.loads(body.decode("utf-8"))
            runtime_log(f"Webhook payload events={len(payload.get('events', []))}")
            host = self.headers.get("X-Forwarded-Host") or self.headers.get("Host", "")
            proto = self.headers.get("X-Forwarded-Proto") or ("https" if "ngrok" in host else "http")
            public_base_url = f"{proto}://{host}" if host else ""
            for event in payload.get("events", []):
                reply = process_line_event_menu(event, public_base_url)
                if reply and event.get("replyToken"):
                    if isinstance(reply, list):
                        reply_messages(event["replyToken"], reply)
                    elif isinstance(reply, dict):
                        reply_messages(event["replyToken"], [reply])
                    else:
                        reply_text(event["replyToken"], reply)
            self._send_json(HTTPStatus.OK, {"status": "ok"})
        except Exception as exc:
            runtime_log(f"LINE webhook error: {exc}")
            self._send_json(HTTPStatus.OK, {"status": "accepted_with_error", "error": str(exc)})

    def log_message(self, format: str, *args: Any) -> None:
        sys.stderr.write("%s - %s\n" % (self.address_string(), format % args))


def main() -> int:
    global CONFIG, CONFIG_PATH, STATE_CACHE
    parser = argparse.ArgumentParser(description="LINE webhook service that imports receipt images into VAT Excel.")
    parser.add_argument("--config", required=True, help="Path to line_bot_config.example.json or your copied config.")
    args = parser.parse_args()

    CONFIG_PATH = Path(args.config).resolve()
    CONFIG = load_config(CONFIG_PATH)
    STATE_CACHE = load_state_cache()
    workbook_path = resolve_path(CONFIG["workbook"])
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    secret_name = CONFIG["line"]["channel_secret_env"]
    token_name = CONFIG["line"]["channel_access_token_env"]
    runtime_log(f"Startup workbook={workbook_path}")
    runtime_log(f"Startup {secret_name} set={bool(os.getenv(secret_name))}")
    runtime_log(f"Startup {token_name} set={bool(os.getenv(token_name))}")
    runtime_log(f"Startup google_apps_script_url set={bool(CONFIG.get('google_apps_script_url'))}")
    runtime_log(f"Startup google_apps_script_secret set={bool(CONFIG.get('google_apps_script_secret'))}")

    host = CONFIG.get("host", "0.0.0.0")
    port = int(CONFIG.get("port", 8080))
    server = ThreadingHTTPServer((host, port), LineWebhookHandler)
    print(f"LINE expense bot listening on http://{host}:{port}/callback")
    print(f"Workbook: {workbook_path}")
    server.serve_forever()
    return 0


if __name__ == "__main__":
    sys.exit(main())
